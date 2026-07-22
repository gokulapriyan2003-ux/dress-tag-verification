import streamlit as st
import pandas as pd
import openpyxl
import os
import sys
import urllib.request
from openpyxl.styles import PatternFill

# Import the core logic from compare_tags.py
sys.path.append(os.path.dirname(os.path.abspath(__file__)))
from compare_tags import (
    extract_pdf_tags,
    extract_excel_master,
    compare,
    get_updated_mrp
)

st.set_page_config(
    page_title="Dress Tag Verifier",
    layout="wide"
)

# Custom Styling for premium aesthetics
st.markdown("""
    <style>
    .main-title {
        font-size: 2.8rem;
        font-weight: 700;
        color: #1E3A8A;
        margin-bottom: 0.1rem;
    }
    .subtitle {
        font-size: 1.2rem;
        color: #4B5563;
        margin-bottom: 2rem;
    }
    .metric-box {
        background-color: #F3F4F6;
        padding: 1.5rem;
        border-radius: 0.75rem;
        text-align: center;
        box-shadow: 0 1px 3px rgba(0,0,0,0.1);
    }
    .metric-value {
        font-size: 2rem;
        font-weight: 700;
        color: #2563EB;
    }
    .metric-label {
        font-size: 0.9rem;
        color: #6B7280;
        text-transform: uppercase;
        letter-spacing: 0.05em;
    }
    </style>
""", unsafe_allow_html=True)

st.markdown('<div class="main-title">Dress Tag & Master Sheet Verifier</div>', unsafe_allow_html=True)
st.markdown('<div class="subtitle">Extract SKU fields from multi-tag PDF and validate them against Excel & Google Sheet references</div>', unsafe_allow_html=True)

# Auto-detect local files
script_dir = os.path.dirname(os.path.abspath(__file__))
local_pdfs = [f for f in os.listdir(script_dir) if f.lower().endswith(".pdf") and not f.startswith("~$")]
local_xlsxs = [f for f in os.listdir(script_dir) if f.lower().endswith(".xlsx") and not f.startswith("~$") and not any(x in f.lower() for x in ["report", "google", "explore", "comparison"])]

default_pdf = os.path.join(script_dir, local_pdfs[0]) if local_pdfs else None
default_xlsx = os.path.join(script_dir, local_xlsxs[0]) if local_xlsxs else None

# Layout: Sidebar configuration
st.sidebar.header("Configuration & Local Files")
st.sidebar.markdown("### Auto-detected files:")
if default_pdf:
    st.sidebar.success(f"PDF found: `{os.path.basename(default_pdf)}`")
else:
    st.sidebar.warning("No local PDF found in directory.")

if default_xlsx:
    st.sidebar.success(f"Excel found: `{os.path.basename(default_xlsx)}`")
else:
    st.sidebar.warning("No local Excel found in directory.")

# Step 1: Tag Verification Mode Selection (Placed BEFORE tag uploading)
st.subheader("1. Select Tag Verification Mode")
tag_type = st.selectbox(
    "Tag Verification Type",
    options=["D2C Dress tag file", "B2B Box Sticker tag file"],
    index=0,
    help="Select 'D2C Dress tag file' for standard dress tags or 'B2B Box Sticker tag file' for B2B box stickers (verifying Lot No, Pack Qty, Total MRP, EAN, SKU)."
)

# Step 2: Reference File Uploaders
st.subheader("2. Upload Reference Files")
col1, col2 = st.columns(2)

with col1:
    pdf_file = st.file_uploader("Upload Tag PDF (Optional, defaults to local file if empty)", type=["pdf"])
with col2:
    xlsx_file = st.file_uploader("Upload Master Excel (Optional, defaults to local file if empty)", type=["xlsx"])

sheet_name = st.sidebar.text_input("Excel Sheet Name (Optional, uses first sheet if blank)", value="")

# Setup paths based on uploads or local defaults
target_pdf = None
if pdf_file:
    # Save uploaded file to temp path
    target_pdf = os.path.join(script_dir, "temp_uploaded_tags.pdf")
    with open(target_pdf, "wb") as f:
        f.write(pdf_file.getbuffer())
elif default_pdf:
    target_pdf = default_pdf

target_xlsx = None
if xlsx_file:
    target_xlsx = os.path.join(script_dir, "temp_uploaded_master.xlsx")
    with open(target_xlsx, "wb") as f:
        f.write(xlsx_file.getbuffer())
elif default_xlsx:
    target_xlsx = default_xlsx

# Run Verification Button
if st.button("Run Verification", type="primary"):
    if not target_pdf:
        st.error("Please upload a PDF file or place one in the script directory.")
    elif not target_xlsx:
        st.error("Please upload a Master Excel file or place one in the script directory.")
    else:
        with st.spinner("Processing..."):
            # Download updated MRP Google Sheet
            gsheet_dfs = {}
            gsheet_url = "https://docs.google.com/spreadsheets/d/1Q7nboN_Rezl807J0naA0QczTyoAQ6WM-KNmp_F26n5M/export?format=xlsx"
            gsheet_path = os.path.join(script_dir, "google_sheet_mrp.xlsx")
            
            try:
                req = urllib.request.Request(gsheet_url, headers={'User-Agent': 'Mozilla/5.0'})
                with urllib.request.urlopen(req, timeout=15) as response:
                    with open(gsheet_path, "wb") as f:
                        f.write(response.read())
                xls = pd.ExcelFile(gsheet_path)
                for name in xls.sheet_names:
                    gsheet_dfs[name] = pd.read_excel(xls, sheet_name=name)
                st.success("Downloaded latest MRP Google Sheet successfully.")
            except Exception as e:
                st.warning(f"Could not download updated MRP Google Sheet ({e}). Falling back to local Excel values.")

            # Load files
            try:
                pdf_df = extract_pdf_tags(target_pdf)
                excel_df = extract_excel_master(target_xlsx, sheet_name if sheet_name else None)
                
                # Perform comparison
                report_df = compare(pdf_df, excel_df, gsheet_dfs, tag_type=tag_type)
                
                n_mismatch = (report_df["Status"] != "✅ Match").sum()
                n_total = len(report_df)
                success_rate = round(((n_total - n_mismatch) / n_total) * 100, 1) if n_total > 0 else 0
                
                # Display Metrics
                st.subheader("Verification Summary")
                m_col1, m_col2, m_col3 = st.columns(3)
                with m_col1:
                    st.markdown(f"""
                        <div class="metric-box">
                            <div class="metric-value">{len(pdf_df)}</div>
                            <div class="metric-label">Tags Extracted</div>
                        </div>
                    """, unsafe_allow_html=True)
                with m_col2:
                    st.markdown(f"""
                        <div class="metric-box">
                            <div class="metric-value">{n_total}</div>
                            <div class="metric-label">Field Checks Run</div>
                        </div>
                    """, unsafe_allow_html=True)
                with m_col3:
                    color = "#10B981" if success_rate == 100 else "#EF4444"
                    st.markdown(f"""
                        <div class="metric-box">
                            <div class="metric-value" style="color: {color}">{success_rate}%</div>
                            <div class="metric-label">Success Rate</div>
                        </div>
                    """, unsafe_allow_html=True)
                
                # Save Report File
                out_path = os.path.join(script_dir, "tag_comparison_report.xlsx")
                try:
                    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
                        pdf_df.to_excel(writer, sheet_name="PDF_Extracted", index=False)
                        excel_df.to_excel(writer, sheet_name="Excel_Master", index=False)
                        report_df.to_excel(writer, sheet_name="Comparison_Report", index=False)

                        # Color the report sheet directly in-memory to prevent BadZipFile errors
                        workbook = writer.book
                        worksheet = writer.sheets["Comparison_Report"]
                        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                        
                        status_col_idx = report_df.columns.get_loc("Status") + 1
                        for row_idx in range(2, len(report_df) + 2):
                            status_val = str(report_df.iloc[row_idx - 2].get("Status", ""))
                            fill = green_fill if "Match" in status_val and "Mis" not in status_val and "Not found" not in status_val else red_fill
                            for col_idx in range(1, len(report_df.columns) + 1):
                                worksheet.cell(row=row_idx, column=col_idx).fill = fill
                except PermissionError:
                    st.error(f"Permission denied when writing to '{out_path}'. Please make sure it is closed in Microsoft Excel and try again.")
                
                # Show results
                if n_mismatch == 0:
                    st.balloons()
                    st.success("All field checks passed successfully! EAN Barcode, Sizes, and MRP values are 100% accurate.")
                else:
                    st.error(f"Found {n_mismatch} mismatches. Please check the report or view details below.")
                    mismatch_df = report_df[report_df["Status"] != "✅ Match"]
                    st.dataframe(mismatch_df, use_container_width=True)

                # Provide Download Link
                if os.path.exists(out_path):
                    with open(out_path, "rb") as file:
                        btn = st.download_button(
                            label="Download Excel Comparison Report",
                            data=file,
                            file_name="tag_comparison_report.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                        )
                
                # Show full comparison table
                with st.expander("View Full Comparison Report Details"):
                    st.dataframe(report_df, use_container_width=True)
                    
            except Exception as ex:
                st.error(f"Error during processing: {ex}")
                st.exception(ex)
