"""
Dress Tag vs Master Sheet Verifier
-----------------------------------
Extracts per-SKU fields from a multi-tag PDF (dress/garment tags laid out
N-per-row) and compares them against a master Excel sheet, producing a
match/mismatch report.

Usage:
    python compare_tags.py <tag_pdf> <master_xlsx> [--sheet SHEET_NAME] [--out report.xlsx]
"""

import os
import re
import sys
import argparse
import pdfplumber
import openpyxl
import pandas as pd
import urllib.request


# ---------------------------------------------------------------------------
# 1. PDF EXTRACTION
# ---------------------------------------------------------------------------

# Each of these labels appears once per tag, repeated N times per printed line
# (N = number of tags side-by-side in that row of the sheet). We split each
# line on the label text itself to recover the N individual values in order.
CANONICAL_LABELS = {
    "Style:": ["Style:", "STYLE:", "Lot No:", "LOT NO:", "STYLE CODE:", "Style Code:", "LOT NO :", "LOT NO  :"],
    "Product:": ["Product:", "PRODUCT:", "Product Name:", "PRODUCT NAME:"],
    "Fit:": ["Fit:", "FIT:"],
    "Color:": ["Color:", "COLOR:", "Colour:", "COLOUR:"],
    "Category:": ["Category:", "CATEGORY:"],
    "Manufactured On:": ["Manufactured On:", "MANUFACTURED ON:", "MFD :", "MFD:", "MFD ON:", "MFD ON :"],
    "Net Quantity:": ["Net Quantity:", "NET QUANTITY:", "Net Qty:", "NET QTY:"],
    "HSN Code:": ["HSN Code:", "HSN CODE:"],
    "SKU Code:": ["SKU Code:", "SKU CODE:", "SKU:"],
    "SIZE :": ["SIZE :", "SIZE:", "Size:", "Size :"],
    "MRP:": ["MRP:", "MRP :"],
    "Qty:": ["Qty:", "QTY:", "Qty :", "QTY :"],
}

LABELS = []
LABEL_TO_CANONICAL = {}
for canonical, variations in CANONICAL_LABELS.items():
    for var in variations:
        LABELS.append(var)
        LABEL_TO_CANONICAL[var] = canonical

BARCODE_RE = re.compile(r"^\d{8,14}$")          # standalone barcode line
CM_RE = re.compile(r"^\(\d+(\.\d+)?CM\)$")       # e.g. (71.12CM)


def split_repeated_label(line: str, label: str):
    parts = line.split(label)
    parts = [p.strip() for p in parts if p.strip() != ""]
    return parts


def split_repeated_label_case_insensitive(line: str, label: str):
    pattern = re.compile(re.escape(label), re.IGNORECASE)
    parts = pattern.split(line)
    parts = [p.strip() for p in parts if p.strip()]
    return parts


def extract_pdf_tags(pdf_path: str) -> pd.DataFrame:
    field_lists = {lbl: [] for lbl in CANONICAL_LABELS.keys()}
    barcodes = []
    cm_sizes = []
    descriptions = []
    single_mrps = []
    total_mrps = []
    pack_quantities = []
    sku_to_huge_size = {}

    SIZE_SET = {"M", "L", "XL", "LX", "2XL", "LX2", "3XL", "LX3", "4XL", "LX4", "5XL", "LX5", "S", "XS", "SX", "XXL", "LXX", "06UK", "07UK", "08UK", "09UK", "10UK", "11UK", "12UK", "6UK", "7UK", "8UK", "9UK", "08Y", "10Y", "12Y", "14Y", "2Y", "4Y", "6Y", "8Y"}
    
    def parse_vertical_reversed_size(w_text):
        rev = w_text[::-1].strip().upper()
        if rev in ["S/H", "S/F"]:
            return None
        if rev.startswith("Y") and len(rev) >= 2 and rev[1:].isdigit():
            return rev[::-1]
        if "/" in rev:
            base = rev.split("/")[0].strip()
            if base in SIZE_SET:
                return rev
        return None

    def clean_reverse_size(sz):
        s = str(sz).strip().upper()
        if s == "LX": return "XL"
        if s == "LX2": return "2XL"
        if s == "LX3": return "3XL"
        if s == "LX4": return "4XL"
        if s == "LX5": return "5XL"
        if s == "SX": return "XS"
        if s == "LXX": return "XXL"
        return s

    def extract_single_desc_from_repeating(s):
        s = str(s).strip()
        words = s.split()
        n_words = len(words)
        for k in range(2, 9):
            if n_words % k == 0 or n_words > k:
                pattern = " ".join(words[:k])
                re_pattern = r"^(" + re.escape(pattern) + r"\s*)+$"
                if re.match(re_pattern, s):
                    return pattern
        return s

    try:
        with pdfplumber.open(pdf_path) as pdf:
            for page in pdf.pages:
                words = page.extract_words()
                skus = []
                for w in words:
                    text = w["text"].strip()
                    if len(text) >= 8 and text.isupper() and any(c.isdigit() for c in text):
                        skus.append(w)
                sizes = []
                for w in words:
                    text = w["text"].strip().upper()
                    vert = parse_vertical_reversed_size(text)
                    if vert:
                        w_copy = dict(w)
                        w_copy["text"] = vert
                        sizes.append(w_copy)
                    elif text in SIZE_SET:
                        sizes.append(w)
                for sku_w in skus:
                    sku_text = sku_w["text"]
                    sku_x = sku_w["x0"]
                    sku_y = sku_w["top"]
                    
                    best_size = None
                    min_dist = float("inf")
                    for sz_w in sizes:
                        sz_text = sz_w["text"]
                        sz_x = sz_w["x0"]
                        sz_y = sz_w["top"]
                        
                        if sz_x < sku_x:
                            continue
                        y_diff = abs(sz_y - sku_y)
                        if y_diff > 120:
                            continue
                        dist = ((sz_x - sku_x)**2 + y_diff**2)**0.5
                        if dist < min_dist:
                            min_dist = dist
                            best_size = sz_text
                    if best_size:
                        sku_to_huge_size[sku_text] = clean_reverse_size(best_size)
    except Exception:
        pass

    with pdfplumber.open(pdf_path) as pdf:
        for page_num, page in enumerate(pdf.pages):
            text = page.extract_text() or ""
            raw_lines = [l.strip() for l in text.split("\n") if l.strip()]
            
            # Check if this page contains tags (must have barcode or MRP)
            has_tags = False
            for line in raw_lines:
                tokens = line.split()
                if any(BARCODE_RE.match(t) for t in tokens):
                    has_tags = True
                    break
                if "MRP" in line.upper():
                    has_tags = True
                    break
            
            if not has_tags:
                continue
                
            for idx_line, line in enumerate(raw_lines):
                if page_num == 0 and idx_line < 7:
                    continue
                # 1. Check for MRP line containing quantities
                mrp_matches = list(re.finditer(r"₹?\s*([\d,]+\.?\d*)\s*/-\s*\(\s*(\d+)\s*(Nos?|Pcs?)\s*\)", line, re.IGNORECASE))
                if mrp_matches:
                    for m in mrp_matches:
                        price = float(m.group(1).replace(",", ""))
                        qty = int(m.group(2))
                        if qty == 1:
                            single_mrps.append(price)
                        else:
                            total_mrps.append(price)
                            pack_quantities.append(qty)
                    continue

                matches = []
                for lbl in LABELS:
                    pattern = re.compile(re.escape(lbl), re.IGNORECASE)
                    for m in pattern.finditer(line):
                        matches.append((m.start(), m.end(), lbl))
                
                matches.sort(key=lambda x: (x[0], -(x[1] - x[0])))
                filtered_matches = []
                last_end = -1
                for start, end, lbl in matches:
                    if start >= last_end:
                        filtered_matches.append((start, end, lbl))
                        last_end = end
                
                if filtered_matches:
                    has_lot_no = any("LOT NO" in lbl.upper() for _, _, lbl in filtered_matches)
                    if has_lot_no and idx_line > 0:
                        if not (page_num == 0 and idx_line < 7):
                            prev_line = raw_lines[idx_line - 1]
                            lots_count = sum(1 for _, _, lbl in filtered_matches if "LOT NO" in lbl.upper())
                            if lots_count > 0:
                                if not any(re.match(r"^" + re.escape(x), prev_line, re.IGNORECASE) for x in LABELS):
                                    parts = [p.strip() for p in prev_line.split("  ") if p.strip()]
                                    if len(parts) != lots_count:
                                        single_desc = extract_single_desc_from_repeating(prev_line)
                                        parts = [single_desc] * lots_count
                                    descriptions.extend(parts)

                    for i, (start, end, lbl) in enumerate(filtered_matches):
                        canonical = LABEL_TO_CANONICAL[lbl]
                        val_start = end
                        val_end = filtered_matches[i+1][0] if i+1 < len(filtered_matches) else len(line)
                        val_str = line[val_start:val_end].strip()

                        if canonical not in ["Product:", "Description"]:
                            parts = [p.strip() for p in val_str.split("  ") if p.strip()]
                            if not parts:
                                parts = [""]
                            field_lists[canonical].extend(parts)
                        else:
                            field_lists[canonical].append(val_str)
                    continue

                # barcode / cm-size lines contain several space-separated tokens
                tokens = line.split()
                found_barcodes = [t for t in tokens if BARCODE_RE.match(t)]
                if found_barcodes:
                    barcodes.extend(found_barcodes)
                    
                found_cm = [t for t in tokens if CM_RE.match(t)]
                if found_cm:
                    cm_sizes.extend(found_cm)

    counts = {k: len(v) for k, v in field_lists.items()}
    counts["Barcode"] = len(barcodes)
    counts["CM"] = len(cm_sizes)
    n_tags = counts.get("SKU Code:", 0)
    if n_tags is None:
        n_tags = 0

    # Sanity check: every field should appear exactly once per tag.
    mismatched = {k: v for k, v in counts.items() if v != n_tags}
    if mismatched:
        print("WARNING: field counts don't all line up 1:1 with tag count "
              f"({n_tags}). Counts: {counts}", file=sys.stderr)

    rows = []
    for i in range(n_tags):
        def get(lbl):
            lst = field_lists[lbl]
            return lst[i] if i < len(lst) else None

        mrp_raw = get("MRP:")
        mrp_val = None
        if i < len(single_mrps):
            mrp_val = single_mrps[i]
        elif mrp_raw:
            m = re.search(r"[\d,]+\.?\d*", mrp_raw.replace("₹", ""))
            if m:
                mrp_val = float(m.group().replace(",", ""))

        qty_val = None
        if i < len(pack_quantities):
            qty_val = pack_quantities[i]
        else:
            qty_raw = get("Qty:")
            qty_val = int(qty_raw) if qty_raw and qty_raw.isdigit() else qty_raw

        style_raw = get("Style:")
        net_qty_raw = get("Net Quantity:")
        mfd_raw = get("Manufactured On:")
        sku_val = get("SKU Code:")
        desc_val = get("Product:") or (descriptions[i] if i < len(descriptions) else None)
        size_val = get("SIZE :")
        if not size_val and sku_val:
            if sku_val in sku_to_huge_size:
                size_val = sku_to_huge_size[sku_val]
            else:
                _, _, ext_sz = extract_sku_details(sku_val)
                if ext_sz:
                    size_val = format_size_as_tag(ext_sz)

        rows.append({
            "Style": style_raw,
            "Product": desc_val,
            "Description": desc_val,
            "Fit": get("Fit:"),
            "Color": get("Color:"),
            "Category": get("Category:"),
            "Manufactured On": mfd_raw,
            "Net Quantity": net_qty_raw,
            "SKU": sku_val,
            "Size": size_val,
            "Size(CM)": cm_sizes[i] if i < len(cm_sizes) else None,
            "Barcode": barcodes[i] if i < len(barcodes) else None,
            "MRP": mrp_val,
            "Total MRP": total_mrps[i] if i < len(total_mrps) else None,
            "Qty": qty_val,
        })

    return pd.DataFrame(rows)


# ---------------------------------------------------------------------------
# 2. EXCEL MASTER SHEET EXTRACTION
# ---------------------------------------------------------------------------

def extract_excel_master(xlsx_path: str, sheet_name: str = None) -> pd.DataFrame:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    
    # If specific sheet_name requested, try that first; else check all sheets
    if sheet_name and sheet_name in wb.sheetnames:
        candidate_sheets = [sheet_name]
    else:
        # Prefer sheets whose names suggest master data: 'MRP', 'MASTER', 'DATA', 'GS1', 'SHEET1'
        sorted_sheets = sorted(
            wb.sheetnames,
            key=lambda s: 0 if any(k in s.upper() for k in ["MRP", "MASTER", "GS1", "DATA", "PRODUCT", "SKU"]) else (1 if "SHEET1" in s.upper() else 2)
        )
        candidate_sheets = sorted_sheets

    header_keywords = [
        "SKU", "ITEM CODE", "ITEM_CODE", "ITEM NO", "ITEM NUMBER", "ITEM",
        "PRODUCT CODE", "MATERIAL", "ARTICLE", "GTIN", "BARCODE", "BAR CODE",
        "STYLE", "PRODUCT NAME", "DESCRIPTION", "MRP"
    ]

    best_sheet = None
    best_header_idx = None
    best_ws = None

    # Pass 1: Look for a sheet and row containing "SKU"
    for s_name in candidate_sheets:
        ws = wb[s_name]
        preview_rows = list(ws.iter_rows(values_only=True, max_row=35))
        for i, row in enumerate(preview_rows):
            cells = [str(c).strip().upper() if c is not None else "" for c in row]
            if any("SKU" in c for c in cells):
                best_sheet = s_name
                best_header_idx = i
                best_ws = ws
                break
        if best_sheet:
            break

    # Pass 2: Look for alternative SKU / Inventory keywords (ITEM CODE, BARCODE, GTIN, STYLE, etc.)
    if best_sheet is None:
        for s_name in candidate_sheets:
            ws = wb[s_name]
            preview_rows = list(ws.iter_rows(values_only=True, max_row=35))
            for i, row in enumerate(preview_rows):
                cells = [str(c).strip().upper() if c is not None else "" for c in row]
                matches = sum(1 for c in cells if any(kw in c for kw in header_keywords))
                if matches >= 2:
                    best_sheet = s_name
                    best_header_idx = i
                    best_ws = ws
                    break
            if best_sheet:
                break

    # Pass 3: Fallback to the first sheet with at least 3 non-empty columns in any row
    if best_sheet is None:
        for s_name in candidate_sheets:
            ws = wb[s_name]
            preview_rows = list(ws.iter_rows(values_only=True, max_row=35))
            for i, row in enumerate(preview_rows):
                non_empty = [c for c in row if c is not None and str(c).strip()]
                if len(non_empty) >= 3:
                    best_sheet = s_name
                    best_header_idx = i
                    best_ws = ws
                    break
            if best_sheet:
                break

    if best_ws is None or best_header_idx is None:
        available = ", ".join(wb.sheetnames)
        raise ValueError(f"Could not find any recognizable header row or table data in the Excel workbook. Available sheets: [{available}]. Please specify the sheet name or check your file format.")

    all_rows = list(best_ws.iter_rows(values_only=True))
    header_raw = all_rows[best_header_idx]
    
    clean_header = []
    seen = {}
    for idx, c in enumerate(header_raw):
        col_name = str(c).strip() if c is not None else ""
        if not col_name:
            col_name = f"Column_{idx+1}"
        if col_name in seen:
            seen[col_name] += 1
            col_name = f"{col_name}_{seen[col_name]}"
        else:
            seen[col_name] = 0
        clean_header.append(col_name)

    data_rows = []
    for row in all_rows[best_header_idx + 1:]:
        if not row or all(c is None or str(c).strip() == "" for c in row):
            continue
        first_cell = str(row[0]).strip().upper() if row[0] is not None else ""
        if first_cell == "TOTAL":
            break
        padded_row = list(row[:len(clean_header)]) + [None] * max(0, len(clean_header) - len(row))
        data_rows.append(padded_row)

    df = pd.DataFrame(data_rows, columns=clean_header)
    return df


# ---------------------------------------------------------------------------
# 3. COMPARISON
# ---------------------------------------------------------------------------

def normalize_sku(x):
    if x is None:
        return ""
    return str(x).strip().upper()


def normalize_lot(x):
    if x is None:
        return ""
    return str(x).strip().upper()


def normalize_text(x):
    if x is None:
        return ""
    import re
    s = str(x).strip().upper()
    s = s.replace("-", " ").replace("/", " ").replace("_", " ")
    
    # Replace compound words and remove gender prefixes
    replacements = {
        "TRACKPANT": "TRACK PANT",
        "TRACKPANTS": "TRACK PANT",
        "TRACK PANTS": "TRACK PANT",
        "TSHIRT": "T SHIRT",
        "TSHIRTS": "T SHIRT",
        "PANTS": "PANT",
        "JOGGERS": "JOGGER",
        "SHORTS": "SHORT",
        "TIGHTS": "TIGHT",
        "WOMENS": "",
        "MENS": "",
        "KIDS": "",
        "BOYS": "",
        "GIRLS": ""
    }
    for old, new in replacements.items():
        s = re.sub(r"\b" + re.escape(old) + r"\b", new, s)
        
    s = re.sub(r"\s+", " ", s).strip()
    return s


def normalize_size(x):
    if x is None:
        return ""
    s = str(x).strip().upper()

    kids_map = {
        "08Y": "30",
        "8Y": "30",
        "10Y": "32",
        "12Y": "34",
        "14Y": "36",
        "06Y": "28",
        "6Y": "28",
        "04Y": "26",
        "4Y": "26",
        "02Y": "24",
        "2Y": "24",
    }
    s = kids_map.get(s, s)

    if "/" in s:
        s = s.split("/")[0].strip()

    if s.endswith("UK"):
        s = s[:-2].strip()

    if (s.startswith("K") or s.startswith("Y")) and len(s) >= 2 and s[1:].isdigit():
        s = s[1:]

    if s.isdigit():
        s = str(int(s))

    if len(s) == 3 and s[0].isalpha() and s[1:].isdigit():
        s = str(int(s[1:]))

    size_map = {
        "XSML": "XS",
        "SML": "S",
        "SMALL": "S",
        "MED": "M",
        "MEDIUM": "M",
        "LAR": "L",
        "LARGE": "L",
        "XLR": "XL",
        "EXTRA LARGE": "XL",
        "2XLR": "2XL",
        "XXL": "2XL",
        "2XL": "2XL",
        "3XLR": "3XL",
        "XXXL": "3XL",
        "3XL": "3XL",
        "4XLR": "4XL",
        "XXXXL": "4XL",
        "4XL": "4XL",
        "5XLR": "5XL",
        "5XL": "5XL",
    }
    return size_map.get(s, s)


def format_size_as_tag(size_str):
    if not size_str:
        return size_str
    s = str(size_str).strip().upper()

    if (s.startswith("K") or s.startswith("Y")) and len(s) >= 2 and s[1:].isdigit():
        digits_str = s[1:]
        if len(digits_str) == 1:
            digits_str = "0" + digits_str
        return digits_str + "UK"

    if len(s) == 3 and s[0].isalpha() and s[1:].isdigit():
        return str(int(s[1:]))

    size_map = {
        "XSML": "XS",
        "SML": "S",
        "MED": "M",
        "LAR": "L",
        "XLR": "XL",
        "2XLR": "2XL",
        "XXL": "2XL",
        "3XLR": "3XL",
        "XXXL": "3XL",
        "4XLR": "4XL",
        "XXXXL": "4XL",
        "5XLR": "5XL",
    }
    return size_map.get(s, s)


color_map = {
    "AAE": "ALLURING AZURE",
    "ALLURING AZURE": "ALLURING AZURE",
    "ABN": "ANT BATTALION",
    "ANT BATTALION": "ANT BATTALION",
    "ABT": "AMERICAN BEAUTY",
    "AMERICAN BEAUTY": "AMERICAN BEAUTY",
    "ACL": "ACID LIME",
    "ACID LIME": "ACID LIME",
    "ACO": "AQUA CAMO",
    "AQUA CAMO": "AQUA CAMO",
    "ADP": "ATLANTIC DEEP",
    "ATLANTIC DEEP": "ATLANTIC DEEP",
    "AGC": "AQUA GREY_GULF COAST B",
    "AQUA GREY_GULF COAST B": "AQUA GREY_GULF COAST B",
    "AGM": "AQUA GREY MELANGE",
    "AQUA GREY MELANGE": "AQUA GREY MELANGE",
    "AGN": "ASSURE GREEN",
    "ASSURE GREEN": "ASSURE GREEN",
    "AGW": "AZURE GLOW",
    "AZURE GLOW": "AZURE GLOW",
    "AGY": "AQUA GREY",
    "AQUA GREY": "AQUA GREY",
    "AHM": "ASH MELANGE",
    "ASH MELANGE": "ASH MELANGE",
    "AME": "AURORA MELANGE",
    "AURORA MELANGE": "AURORA MELANGE",
    "AQA": "AQUA",
    "AQUA": "AQUA",
    "AQM": "AQUA MELANGE",
    "AQUA MELANGE": "AQUA MELANGE",
    "AQO": "AQUA OCEAN",
    "AQUA OCEAN": "AQUA OCEAN",
    "AQS": "AQUA SEA",
    "AQUA SEA": "AQUA SEA",
    "ARA": "AURORA",
    "AURORA": "AURORA",
    "ASH": "ASH",
    "ASH": "ASH",
    "ASK": "ABSTRACT STREAK",
    "ABSTRACT STREAK": "ABSTRACT STREAK",
    "ASL": "AQUA SWIRL",
    "AQUA SWIRL": "AQUA SWIRL",
    "ASM": "AQUA SEA MELANGE",
    "AQUA SEA MELANGE": "AQUA SEA MELANGE",
    "ATY": "ALLOTROPHY",
    "ALLOTROPHY": "ALLOTROPHY",
    "AWG": "ASH_WOODLAND GRAY",
    "ASH_WOODLAND GRAY": "ASH_WOODLAND GRAY",
    "BAM": "BLACK MYRTLE",
    "BLACK MYRTLE": "BLACK MYRTLE",
    "BAS": "BLUE AMORPHOUS",
    "BLUE AMORPHOUS": "BLUE AMORPHOUS",
    "BBD": "BLACK NEO_BUILT DIFFERENT",
    "BLACK NEO_BUILT DIFFERENT": "BLACK NEO_BUILT DIFFERENT",
    "BBE": "BRUSHED BLUE",
    "BRUSHED BLUE": "BRUSHED BLUE",
    "BBH": "BROWN BRUSH",
    "BROWN BRUSH": "BROWN BRUSH",
    "BBL": "BLACK BASKET BALL",
    "BLACK BASKET BALL": "BLACK BASKET BALL",
    "BBR": "BLACK_BERRY RED",
    "BLACK_BERRY RED": "BLACK_BERRY RED",
    "BBT": "BLACK BLUE TIDE",
    "BLACK BLUE TIDE": "BLACK BLUE TIDE",
    "BBV": "BLUE BERRY VIOLET",
    "BLUE BERRY VIOLET": "BLUE BERRY VIOLET",
    "BCG": "BLACK CARBON GREY",
    "BLACK CARBON GREY": "BLACK CARBON GREY",
    "BCH": "BLACK_CHOCOLATE PRO",
    "BLACK_CHOCOLATE PRO": "BLACK_CHOCOLATE PRO",
    "BCL": "BLUE CORAL",
    "BLUE CORAL": "BLUE CORAL",
    "BCM": "BLACK CURRENT MELANGE",
    "BLACK CURRENT MELANGE": "BLACK CURRENT MELANGE",
    "BCO": "BLACK CAMO",
    "BLACK CAMO": "BLACK CAMO",
    "BCS": "BLACK CANVAS",
    "BLACK CANVAS": "BLACK CANVAS",
    "BCT": "BLACK CURRENT",
    "BLACK CURRENT": "BLACK CURRENT",
    "BCY": "BLACK CITY GREY",
    "BLACK CITY GREY": "BLACK CITY GREY",
    "BDA": "BLUE DELTA",
    "BLUE DELTA": "BLUE DELTA",
    "BEB": "BLUE BOSCAGE",
    "BLUE BOSCAGE": "BLUE BOSCAGE",
    "BEC": "BLUE CAMO",
    "BLUE CAMO": "BLUE CAMO",
    "BEG": "BLUE GENIE",
    "BLUE GENIE": "BLUE GENIE",
    "BEK": "BUBBLE PINK",
    "BUBBLE PINK": "BUBBLE PINK",
    "BEL": "BLUE TEAL",
    "BLUE TEAL": "BLUE TEAL",
    "BEM": "BLUE TIDE MELANGE",
    "BLUE TIDE MELANGE": "BLUE TIDE MELANGE",
    "BES": "BLUE STRIPES",
    "BLUE STRIPES": "BLUE STRIPES",
    "BET": "BLUE LIGHT MELANGE",
    "BLUE LIGHT MELANGE": "BLUE LIGHT MELANGE",
    "BFM": "BAY LEAF MELANGE",
    "BAY LEAF MELANGE": "BAY LEAF MELANGE",
    "BGE": "BLUE GRANITE",
    "BLUE GRANITE": "BLUE GRANITE",
    "BGM": "BURGUNDY MELANGE",
    "BURGUNDY MELANGE": "BURGUNDY MELANGE",
    "BGT": "BLACK_GULF COAST B",
    "BLACK_GULF COAST B": "BLACK_GULF COAST B",
    "BGY": "BLUE GREY",
    "BLUE GREY": "BLUE GREY",
    "BHG": "BLACK HUNTER GREEN",
    "BLACK HUNTER GREEN": "BLACK HUNTER GREEN",
    "BHS": "BRUSH STROKES",
    "BRUSH STROKES": "BRUSH STROKES",
    "BIG": "BLACK IRON GREY",
    "BLACK IRON GREY": "BLACK IRON GREY",
    "BKA": "BLACK ASCENT",
    "BLACK ASCENT": "BLACK ASCENT",
    "BKB": "BLACK BELIEVE",
    "BLACK BELIEVE": "BLACK BELIEVE",
    "BKD": "BLACK DIAMOND",
    "BLACK DIAMOND": "BLACK DIAMOND",
    "BKE": "BLACK MIRAGE",
    "BLACK MIRAGE": "BLACK MIRAGE",
    "BKF": "BLACK FORWARD",
    "BLACK FORWARD": "BLACK FORWARD",
    "BKG": "BLACK GRUNGE",
    "BLACK GRUNGE": "BLACK GRUNGE",
    "BKM": "BLACK MAGIC",
    "BLACK MAGIC": "BLACK MAGIC",
    "BKR": "BLACK RED",
    "BLACK RED": "BLACK RED",
    "BKS": "BLACK SHUTTLE",
    "BLACK SHUTTLE": "BLACK SHUTTLE",
    "BKW": "BLACKY LUMIWAVE",
    "BLACKY LUMIWAVE": "BLACKY LUMIWAVE",
    "BLA": "BLACK ASTRONAUT",
    "BLACK ASTRONAUT": "BLACK ASTRONAUT",
    "BLB": "BLACK BEETLE",
    "BLACK BEETLE": "BLACK BEETLE",
    "BLC": "BLACK NEO_LOCKED IT",
    "BLACK NEO_LOCKED IT": "BLACK NEO_LOCKED IT",
    "BLD": "BLACK DR CARBON",
    "BLACK DR CARBON": "BLACK DR CARBON",
    "BLE": "BLUE STONE",
    "BLUE STONE": "BLUE STONE",
    "BLEX": "BLACK EXPLORE",
    "BLACK EXPLORE": "BLACK EXPLORE",
    "BLF": "NO",
    "NO": "NO",
    "BLG": "BLACK-LT GREY",
    "BLACK-LT GREY": "BLACK-LT GREY",
    "BLGR": "BLACK GARP",
    "BLACK GARP": "BLACK GARP",
    "BLH": "BLUE LABYRINTH",
    "BLUE LABYRINTH": "BLUE LABYRINTH",
    "BLI": "BLACK LINDEN",
    "BLACK LINDEN": "BLACK LINDEN",
    "BLK": "BLACK",
    "BLACK": "BLACK",
    "BLL": "BLACK LUNAR GREY",
    "BLACK LUNAR GREY": "BLACK LUNAR GREY",
    "BLM": "BLACK MELANGE",
    "BLACK MELANGE": "BLACK MELANGE",
    "BLP": "BLUE LEOPARD",
    "BLUE LEOPARD": "BLUE LEOPARD",
    "BLS": "BLACK STONE GREY",
    "BLACK STONE GREY": "BLACK STONE GREY",
    "BLT": "BLUE LIGHT",
    "BLUE LIGHT": "BLUE LIGHT",
    "BLW": "BLUEY LUMIWAVE",
    "BLUEY LUMIWAVE": "BLUEY LUMIWAVE",
    "BMA": "BLACK MARSALA",
    "BLACK MARSALA": "BLACK MARSALA",
    "BMB": "BOMBAY BLUE",
    "BOMBAY BLUE": "BOMBAY BLUE",
    "BMD": "BREAD MOULD",
    "BREAD MOULD": "BREAD MOULD",
    "BME": "BEETLE MELANGE",
    "BEETLE MELANGE": "BEETLE MELANGE",
    "BMM": "BLUE STONE MELANGE",
    "BLUE STONE MELANGE": "BLUE STONE MELANGE",
    "BMN": "BADMINTON",
    "BADMINTON": "BADMINTON",
    "BNM": "BROWN MELANGE",
    "BROWN MELANGE": "BROWN MELANGE",
    "BNS": "BROWN STRIPES",
    "BROWN STRIPES": "BROWN STRIPES",
    "BOE": "BLACK OLIVE",
    "BLACK OLIVE": "BLACK OLIVE",
    "BOM": "BLACK OLIVE MELANGE",
    "BLACK OLIVE MELANGE": "BLACK OLIVE MELANGE",
    "BPK": "BABY PINK",
    "BABY PINK": "BABY PINK",
    "BRD": "BERRY RED",
    "BERRY RED": "BERRY RED",
    "BRE": "BLUE MIRAGE",
    "BLUE MIRAGE": "BLUE MIRAGE",
    "BRM": "BERRY RED MELANGE",
    "BERRY RED MELANGE": "BERRY RED MELANGE",
    "BRN": "BROWN",
    "BROWN": "BROWN",
    "BRR": "BLUEY RADIANT RIPPLE",
    "BLUEY RADIANT RIPPLE": "BLUEY RADIANT RIPPLE",
    "BSE": "BLACK STROKE",
    "BLACK STROKE": "BLACK STROKE",
    "BSL": "BLUE SWIRL",
    "BLUE SWIRL": "BLUE SWIRL",
    "BSM": "BLACK STORM",
    "BLACK STORM": "BLACK STORM",
    "BSO": "BLACK_SMOKEY OLIVE",
    "BLACK_SMOKEY OLIVE": "BLACK_SMOKEY OLIVE",
    "BSS": "BLUE STROKES",
    "BLUE STROKES": "BLUE STROKES",
    "BST": "BLUESTONE",
    "BLUESTONE": "BLUESTONE",
    "BSU": "BLACK SUMMIT",
    "BLACK SUMMIT": "BLACK SUMMIT",
    "BTD": "BLUE TIDE",
    "BLUE TIDE": "BLUE TIDE",
    "BTE": "BEETLE",
    "BEETLE": "BEETLE",
    "BTG": "BLACK_TREKKING GREEN",
    "BLACK_TREKKING GREEN": "BLACK_TREKKING GREEN",
    "BTH": "BLUE TURKISH",
    "BLUE TURKISH": "BLUE TURKISH",
    "BTL": "BRIGHT TEAL",
    "BRIGHT TEAL": "BRIGHT TEAL",
    "BTM": "BRIGHT TEAL MELANGE",
    "BRIGHT TEAL MELANGE": "BRIGHT TEAL MELANGE",
    "BTQ": "BLUE TURQUOISE",
    "BLUE TURQUOISE": "BLUE TURQUOISE",
    "BUG": "BURGUNDY",
    "BURGUNDY": "BURGUNDY",
    "BVA": "BOUGAINVILLEA",
    "BOUGAINVILLEA": "BOUGAINVILLEA",
    "BVM": "BLUE BERRY VIOLET MELANGE",
    "BLUE BERRY VIOLET MELANGE": "BLUE BERRY VIOLET MELANGE",
    "BWG": "BLACK_WOODLAND GRAY",
    "BLACK_WOODLAND GRAY": "BLACK_WOODLAND GRAY",
    "BWS": "BLUE WAVES",
    "BLUE WAVES": "BLUE WAVES",
    "CBE": "COBALT BLUE",
    "COBALT BLUE": "COBALT BLUE",
    "CBH": "CERULEAN BLUSH",
    "CERULEAN BLUSH": "CERULEAN BLUSH",
    "CBK": "CAT BLACK",
    "CAT BLACK": "CAT BLACK",
    "CBL": "CARBON_BLACK",
    "CARBON_BLACK": "CARBON_BLACK",
    "CBM": "CARBON MELANGE",
    "CARBON MELANGE": "CARBON MELANGE",
    "CBN": "CARBON",
    "CARBON": "CARBON",
    "CBT": "CLOUD BURST",
    "CLOUD BURST": "CLOUD BURST",
    "CCD": "CAMO CLOUD",
    "CAMO CLOUD": "CAMO CLOUD",
    "CCM": "COCOA CREAM",
    "COCOA CREAM": "COCOA CREAM",
    "CCO": "CARBON CAMO",
    "CARBON CAMO": "CARBON CAMO",
    "CCY": "CAT CLAY",
    "CAT CLAY": "CAT CLAY",
    "CEB": "CODE BLUE",
    "CODE BLUE": "CODE BLUE",
    "CEM": "CAPRI BLUE MELANGE",
    "CAPRI BLUE MELANGE": "CAPRI BLUE MELANGE",
    "CFB": "COFFEE BEAN",
    "COFFEE BEAN": "COFFEE BEAN",
    "CFK": "CHILI FLAKES",
    "CHILI FLAKES": "CHILI FLAKES",
    "CFM": "CHILI FLAKES MELANGE",
    "CHILI FLAKES MELANGE": "CHILI FLAKES MELANGE",
    "CGE": "COGNAE",
    "COGNAE": "COGNAE",
    "CGM": "CARBON GREY MELANGE",
    "CARBON GREY MELANGE": "CARBON GREY MELANGE",
    "CGY": "CITY GREY",
    "CITY GREY": "CITY GREY",
    "CHA": "CHOCOLATE A",
    "CHOCOLATE A": "CHOCOLATE A",
    "CHS": "CHRONOS SQUARE",
    "CHRONOS SQUARE": "CHRONOS SQUARE",
    "CIM": "CHILLI OIL MIL",
    "CHILLI OIL MIL": "CHILLI OIL MIL",
    "CLB": "COOL BLUE",
    "COOL BLUE": "COOL BLUE",
    "CLM": "CAT LIGHT GREY MELANGE",
    "CAT LIGHT GREY MELANGE": "CAT LIGHT GREY MELANGE",
    "CLS": "CLAWS",
    "CLAWS": "CLAWS",
    "CLY": "CLAY",
    "CLAY": "CLAY",
    "CME": "CHOCOLATE  MELANGE",
    "CHOCOLATE  MELANGE": "CHOCOLATE  MELANGE",
    "CMT": "COOL MINT",
    "COOL MINT": "COOL MINT",
    "CNA": "CARBON GREY ASH",
    "CARBON GREY ASH": "CARBON GREY ASH",
    "CNG": "CARBON GREY",
    "CARBON GREY": "CARBON GREY",
    "CNM": "COFFEE BEAN MELANGE",
    "COFFEE BEAN MELANGE": "COFFEE BEAN MELANGE",
    "CNR": "CARBON GREY GRAY RIDGE",
    "CARBON GREY GRAY RIDGE": "CARBON GREY GRAY RIDGE",
    "COL": "CHILLI OIL",
    "CHILLI OIL": "CHILLI OIL",
    "COM": "CHILLI OIL MELANGE",
    "CHILLI OIL MELANGE": "CHILLI OIL MELANGE",
    "COT": "COBALT",
    "COBALT": "COBALT",
    "CRB": "CAPRI BLUE",
    "CAPRI BLUE": "CAPRI BLUE",
    "CRD": "CORAL RED",
    "CORAL RED": "CORAL RED",
    "CRM": "CORAL RED-MELANGE",
    "CORAL RED-MELANGE": "CORAL RED-MELANGE",
    "CSD": "CROSSROAD",
    "CROSSROAD": "CROSSROAD",
    "CSE": "CLOUDSCAPE",
    "CLOUDSCAPE": "CLOUDSCAPE",
    "CSH": "CAMO SPLASH",
    "CAMO SPLASH": "CAMO SPLASH",
    "CSN": "CRIMSON",
    "CRIMSON": "CRIMSON",
    "CST": "CYCLE SPORT",
    "CYCLE SPORT": "CYCLE SPORT",
    "CTA": "CAT AQUA",
    "CAT AQUA": "CAT AQUA",
    "CTF": "CHOCOLATE TRUFFLE",
    "CHOCOLATE TRUFFLE": "CHOCOLATE TRUFFLE",
    "CTG": "CAT LIGHT GREY",
    "CAT LIGHT GREY": "CAT LIGHT GREY",
    "CTL": "CAT LIGHT LIME",
    "CAT LIGHT LIME": "CAT LIGHT LIME",
    "CTM": "CHOCOLATE TRUFFLE MELANGE",
    "CHOCOLATE TRUFFLE MELANGE": "CHOCOLATE TRUFFLE MELANGE",
    "CTN": "CAT NAVY",
    "CAT NAVY": "CAT NAVY",
    "CTO": "CAT ONION",
    "CAT ONION": "CAT ONION",
    "CTR": "CAT RED",
    "CAT RED": "CAT RED",
    "CTS": "COBALT SQUADRON",
    "COBALT SQUADRON": "COBALT SQUADRON",
    "CTT": "CAT LIGHT TURKISH",
    "CAT LIGHT TURKISH": "CAT LIGHT TURKISH",
    "CTV": "CAT BLUEBERRY VIOLET",
    "CAT BLUEBERRY VIOLET": "CAT BLUEBERRY VIOLET",
    "CWA": "CAT WATER AQUA",
    "CAT WATER AQUA": "CAT WATER AQUA",
    "CWB": "COB WEB",
    "COB WEB": "COB WEB",
    "CYM": "CLAY MELANGE",
    "CLAY MELANGE": "CLAY MELANGE",
    "CYW": "CASTRO YELLOW",
    "CASTRO YELLOW": "CASTRO YELLOW",
    "DBE": "DARK BLUE",
    "DARK BLUE": "DARK BLUE",
    "DCC": "DARK CARBON CAMO",
    "DARK CARBON CAMO": "DARK CARBON CAMO",
    "DCH": "DAISY CHAIN",
    "DAISY CHAIN": "DAISY CHAIN",
    "DCM": "DARK CARBON MELANGE",
    "DARK CARBON MELANGE": "DARK CARBON MELANGE",
    "DCN": "DARK CARBON",
    "DARK CARBON": "DARK CARBON",
    "DCO": "DESERT CAMO",
    "DESERT CAMO": "DESERT CAMO",
    "DCS": "DOUBLE CROSS",
    "DOUBLE CROSS": "DOUBLE CROSS",
    "DDC": "DUSTED CAMO",
    "DUSTED CAMO": "DUSTED CAMO",
    "DFM": "DEEP FOREST MELANGE",
    "DEEP FOREST MELANGE": "DEEP FOREST MELANGE",
    "DFS": "DARK FORTRESS",
    "DARK FORTRESS": "DARK FORTRESS",
    "DFT": "DEEP FOREST",
    "DEEP FOREST": "DEEP FOREST",
    "DGY": "DARK GREY",
    "DARK GREY": "DARK GREY",
    "DHE": "DUAL HUE",
    "DUAL HUE": "DUAL HUE",
    "DKE": "DARK EARTH",
    "DARK EARTH": "DARK EARTH",
    "DKH": "DARK HUSH",
    "DARK HUSH": "DARK HUSH",
    "DKR": "DARK RED",
    "DARK RED": "DARK RED",
    "DLH": "DARK LABYRINTH",
    "DARK LABYRINTH": "DARK LABYRINTH",
    "DMM": "DENIM MELANGE",
    "DENIM MELANGE": "DENIM MELANGE",
    "DMR": "DARK MATTER",
    "DARK MATTER": "DARK MATTER",
    "DMT": "DARK MIST",
    "DARK MIST": "DARK MIST",
    "DNM": "DENIM",
    "DENIM": "DENIM",
    "DPE": "DARK PURPLE",
    "DARK PURPLE": "DARK PURPLE",
    "DPK": "DUSTY PINK",
    "DUSTY PINK": "DUSTY PINK",
    "DPM": "DUSTY PINK MELANGE",
    "DUSTY PINK MELANGE": "DUSTY PINK MELANGE",
    "DRD": "DUSTY RED",
    "DUSTY RED": "DUSTY RED",
    "DRE": "DUSTY ROSE",
    "DUSTY ROSE": "DUSTY ROSE",
    "DRM": "DUSTY ROSE MELANGE",
    "DUSTY ROSE MELANGE": "DUSTY ROSE MELANGE",
    "DSB": "DRESS BLUE",
    "DRESS BLUE": "DRESS BLUE",
    "DSK": "DUSK",
    "DUSK": "DUSK",
    "DSN": "DAMSON",
    "DAMSON": "DAMSON",
    "DTE": "DEEP TAUPE",
    "DEEP TAUPE": "DEEP TAUPE",
    "DTM": "DEEP TAUPE MELANGE",
    "DEEP TAUPE MELANGE": "DEEP TAUPE MELANGE",
    "DTY": "DARK TEAL",
    "DARK TEAL": "DARK TEAL",
    "DYM": "DUSTY RED MELANGE",
    "DUSTY RED MELANGE": "DUSTY RED MELANGE",
    "EBE": "EVENING BLUE",
    "EVENING BLUE": "EVENING BLUE",
    "EBM": "EVENING BLUE MELANGE",
    "EVENING BLUE MELANGE": "EVENING BLUE MELANGE",
    "EEG": "EASTER EGG",
    "EASTER EGG": "EASTER EGG",
    "EET": "ECHO ELEMENT",
    "ECHO ELEMENT": "ECHO ELEMENT",
    "ESE": "EBONY SMUDGE",
    "EBONY SMUDGE": "EBONY SMUDGE",
    "ESY": "ENDLESS SKY",
    "ENDLESS SKY": "ENDLESS SKY",
    "FBE": "FUMY BLUE",
    "FUMY BLUE": "FUMY BLUE",
    "FCS": "FOREST CUBES",
    "FOREST CUBES": "FOREST CUBES",
    "FCT": "FOOTBALL COURT",
    "FOOTBALL COURT": "FOOTBALL COURT",
    "FCY": "FIERY CANOPY",
    "FIERY CANOPY": "FIERY CANOPY",
    "FGM": "FIG MELANGE",
    "FIG MELANGE": "FIG MELANGE",
    "FGV": "FOX GLOVE",
    "FOX GLOVE": "FOX GLOVE",
    "FIG": "FIG",
    "FIG": "FIG",
    "FNS": "FLATTEN STONE",
    "FLATTEN STONE": "FLATTEN STONE",
    "FPE": "FUSHIA PURPLE",
    "FUSHIA PURPLE": "FUSHIA PURPLE",
    "FSL": "FREE STYLE LITE",
    "FREE STYLE LITE": "FREE STYLE LITE",
    "FST": "FOOTBALL SPORT",
    "FOOTBALL SPORT": "FOOTBALL SPORT",
    "FWN": "FAWN",
    "FAWN": "FAWN",
    "FXM": "FOX GLOVE MELANGE",
    "FOX GLOVE MELANGE": "FOX GLOVE MELANGE",
    "GBE": "GREY BOSCAGE",
    "GREY BOSCAGE": "GREY BOSCAGE",
    "GBH": "GREEN BRUSH",
    "GREEN BRUSH": "GREEN BRUSH",
    "GCE": "GRUNGE CUBE",
    "GRUNGE CUBE": "GRUNGE CUBE",
    "GCM": "GULF COAST MELANGE",
    "GULF COAST MELANGE": "GULF COAST MELANGE",
    "GCO": "GREY CAMO",
    "GREY CAMO": "GREY CAMO",
    "GDA": "GREEN DELTA",
    "GREEN DELTA": "GREEN DELTA",
    "GFT": "GULF COAST",
    "GULF COAST": "GULF COAST",
    "GGH": "GREEN GLITCH",
    "GREEN GLITCH": "GREEN GLITCH",
    "GHE": "GOLDEN HAZE",
    "GOLDEN HAZE": "GOLDEN HAZE",
    "GJD": "GRAYED JADE",
    "GRAYED JADE": "GRAYED JADE",
    "GLB": "GLACIAL BLUE",
    "GLACIAL BLUE": "GLACIAL BLUE",
    "GLS": "GRASSLANDS",
    "GRASSLANDS": "GRASSLANDS",
    "GME": "GREY MARBLE",
    "GREY MARBLE": "GREY MARBLE",
    "GMI": "GREY MIST",
    "GREY MIST": "GREY MIST",
    "GMR": "GREY MATTER",
    "GREY MATTER": "GREY MATTER",
    "GMT": "GLACIER MELT",
    "GLACIER MELT": "GLACIER MELT",
    "GNB": "GREEN BOG",
    "GREEN BOG": "GREEN BOG",
    "GNM": "GREEN MIST",
    "GREEN MIST": "GREEN MIST",
    "GNS": "GREEN SMUDGE",
    "GREEN SMUDGE": "GREEN SMUDGE",
    "GPE": "GREY PLAGUE",
    "GREY PLAGUE": "GREY PLAGUE",
    "GPN": "GREEN PLATOON",
    "GREEN PLATOON": "GREEN PLATOON",
    "GRM": "GRAY RIDGE MELANGE",
    "GRAY RIDGE MELANGE": "GRAY RIDGE MELANGE",
    "GRR": "GREENY RADIANT RIPPLE",
    "GREENY RADIANT RIPPLE": "GREENY RADIANT RIPPLE",
    "GRY": "GREY",
    "GREY": "GREY",
    "GSE": "GREEN SMOKE",
    "GREEN SMOKE": "GREEN SMOKE",
    "GSL": "GREY SWIRL",
    "GREY SWIRL": "GREY SWIRL",
    "GSS": "GREEN STRIPES",
    "GREEN STRIPES": "GREEN STRIPES",
    "GST": "GREEN STROKES",
    "GREEN STROKES": "GREEN STROKES",
    "GTE": "GREEN TRACE",
    "GREEN TRACE": "GREEN TRACE",
    "GTH": "GLITCH",
    "GLITCH": "GLITCH",
    "GUL": "GULL",
    "GULL": "GULL",
    "GVE": "GRAPE VINE",
    "GRAPE VINE": "GRAPE VINE",
    "GVM": "GRAPE VINE MELANGE",
    "GRAPE VINE MELANGE": "GRAPE VINE MELANGE",
    "GWE": "GREY WAVE",
    "GREY WAVE": "GREY WAVE",
    "GYB": "GREY BLUE",
    "GREY BLUE": "GREY BLUE",
    "GYE": "GREY STROKE",
    "GREY STROKE": "GREY STROKE",
    "GYM": "GREY MELANGE",
    "GREY MELANGE": "GREY MELANGE",
    "GYR": "GRAY RIDGE",
    "GRAY RIDGE": "GRAY RIDGE",
    "GYS": "GREY SMUDGE",
    "GREY SMUDGE": "GREY SMUDGE",
    "HBE": "HORIZON BLUE",
    "HORIZON BLUE": "HORIZON BLUE",
    "HBM": "HORIZON BLUE MELANGE",
    "HORIZON BLUE MELANGE": "HORIZON BLUE MELANGE",
    "HMS": "HUMUS",
    "HUMUS": "HUMUS",
    "HPK": "HOT PINK",
    "HOT PINK": "HOT PINK",
    "HRM": "HEATHER MELANGE",
    "HEATHER MELANGE": "HEATHER MELANGE",
    "HRT": "HORIZONTAL RIFT",
    "HORIZONTAL RIFT": "HORIZONTAL RIFT",
    "HTG": "HUNTER GREEN",
    "HUNTER GREEN": "HUNTER GREEN",
    "HTM": "HUNTER GREEN MELANGE",
    "HUNTER GREEN MELANGE": "HUNTER GREEN MELANGE",
    "HTR": "HEATHER",
    "HEATHER": "HEATHER",
    "IAO": "INDIA ORANGE",
    "INDIA ORANGE": "INDIA ORANGE",
    "IBE": "INK BLUE",
    "INK BLUE": "INK BLUE",
    "IBM": "INDIGO BLOOM",
    "INDIGO BLOOM": "INDIGO BLOOM",
    "IGM": "IRON GREY MELANGE",
    "IRON GREY MELANGE": "IRON GREY MELANGE",
    "IGO": "INDIGO",
    "INDIGO": "INDIGO",
    "IRG": "IRON GREY",
    "IRON GREY": "IRON GREY",
    "ISN": "INK STAIN",
    "INK STAIN": "INK STAIN",
    "JSC": "JIGSAW CAMO",
    "JIGSAW CAMO": "JIGSAW CAMO",
    "JVS": "JUNGLE VINES",
    "JUNGLE VINES": "JUNGLE VINES",
    "KBL": "KHAKI_BLACK",
    "KHAKI_BLACK": "KHAKI_BLACK",
    "KHI": "KHAKI",
    "KHAKI": "KHAKI",
    "KME": "KHAKI MELANGE",
    "KHAKI MELANGE": "KHAKI MELANGE",
    "KPN": "KRYPTON",
    "KRYPTON": "KRYPTON",
    "KUI": "KUI",
    "KUI": "KUI",
    "LAM": "LILAC MELANGE",
    "LILAC MELANGE": "LILAC MELANGE",
    "LBE": "LICHEN BLUE",
    "LICHEN BLUE": "LICHEN BLUE",
    "LBM": "LICHEN BLUE MELANGE",
    "LICHEN BLUE MELANGE": "LICHEN BLUE MELANGE",
    "LCB": "LILAC BREEZE",
    "LILAC BREEZE": "LILAC BREEZE",
    "LCC": "LILAC CAMO",
    "LILAC CAMO": "LILAC CAMO",
    "LCM": "LIGHT CARBON MELANGE",
    "LIGHT CARBON MELANGE": "LIGHT CARBON MELANGE",
    "LDN": "LINDEN",
    "LINDEN": "LINDEN",
    "LEM": "LAKE GREEN MELANGE",
    "LAKE GREEN MELANGE": "LAKE GREEN MELANGE",
    "LFX": "LINEAR FLUX",
    "LINEAR FLUX": "LINEAR FLUX",
    "LGM": "LIGHT GREY MELANGE",
    "LIGHT GREY MELANGE": "LIGHT GREY MELANGE",
    "LGN": "LAKE GREEN",
    "LAKE GREEN": "LAKE GREEN",
    "LHS": "LAYERED HEIGHTS",
    "LAYERED HEIGHTS": "LAYERED HEIGHTS",
    "LLA": "LAVISH LILAC",
    "LAVISH LILAC": "LAVISH LILAC",
    "LLC": "LILAC",
    "LILAC": "LILAC",
    "LLD": "LEGO LAND",
    "LEGO LAND": "LEGO LAND",
    "LLG": "LAUREL GREEN",
    "LAUREL GREEN": "LAUREL GREEN",
    "LLM": "LIGHT LIME MELANGE",
    "LIGHT LIME MELANGE": "LIGHT LIME MELANGE",
    "LME": "LINE MAZE",
    "LINE MAZE": "LINE MAZE",
    "LNB": "LIGHT GREY B_NAVY B",
    "LIGHT GREY B_NAVY B": "LIGHT GREY B_NAVY B",
    "LNM": "LIGHT NAVY MELANGE",
    "LIGHT NAVY MELANGE": "LIGHT NAVY MELANGE",
    "LNR": "LAVENDER",
    "LAVENDER": "LAVENDER",
    "LNS": "LIGHT NAVY SHADOW LIME",
    "LIGHT NAVY SHADOW LIME": "LIGHT NAVY SHADOW LIME",
    "LNT": "LIGHT NAVY TROPOSPHERE",
    "LIGHT NAVY TROPOSPHERE": "LIGHT NAVY TROPOSPHERE",
    "LNW": "LIGHT NAVY WHITE",
    "LIGHT NAVY WHITE": "LIGHT NAVY WHITE",
    "LPD": "LEOPARD",
    "LEOPARD": "LEOPARD",
    "LPE": "LIGHT PURPLE",
    "LIGHT PURPLE": "LIGHT PURPLE",
    "LRG": "LUNAR GREY",
    "LUNAR GREY": "LUNAR GREY",
    "LRN": "LIGHT GREEN",
    "LIGHT GREEN": "LIGHT GREEN",
    "LSE": "LIMESTONE",
    "LIMESTONE": "LIMESTONE",
    "LTB": "LIGHT BLUE",
    "LIGHT BLUE": "LIGHT BLUE",
    "LTC": "LIGHT CARBON",
    "LIGHT CARBON": "LIGHT CARBON",
    "LTG": "LIGHT GREY",
    "LIGHT GREY": "LIGHT GREY",
    "LTL": "LIGHT LIME",
    "LIGHT LIME": "LIGHT LIME",
    "LTN": "LIGHT NAVY",
    "LIGHT NAVY": "LIGHT NAVY",
    "LTO": "LIGHT OLIVE",
    "LIGHT OLIVE": "LIGHT OLIVE",
    "LWG": "LIGHT GREY_WOODLAND GRAY",
    "LIGHT GREY_WOODLAND GRAY": "LIGHT GREY_WOODLAND GRAY",
    "LYM": "LUNAR GREY MELANGE",
    "LUNAR GREY MELANGE": "LUNAR GREY MELANGE",
    "MBE": "MAJOLICA BLUE",
    "MAJOLICA BLUE": "MAJOLICA BLUE",
    "MBG": "MODERN BROWN MELANGE",
    "MODERN BROWN MELANGE": "MODERN BROWN MELANGE",
    "MBM": "MIDNIGHT BLUE MELANGE",
    "MIDNIGHT BLUE MELANGE": "MIDNIGHT BLUE MELANGE",
    "MBN": "MODERN BROWN",
    "MODERN BROWN": "MODERN BROWN",
    "MCS": "MELTING CRAYONS",
    "MELTING CRAYONS": "MELTING CRAYONS",
    "MDG": "MALLARD GREEN",
    "MALLARD GREEN": "MALLARD GREEN",
    "MDM": "MEDIUM DENIM",
    "MEDIUM DENIM": "MEDIUM DENIM",
    "MEM": "MYRTLE MELANGE",
    "MYRTLE MELANGE": "MYRTLE MELANGE",
    "MGM": "MINT GREEN MELANGE",
    "MINT GREEN MELANGE": "MINT GREEN MELANGE",
    "MGN": "MOSS GREEN",
    "MOSS GREEN": "MOSS GREEN",
    "MGO": "MANGO",
    "MANGO": "MANGO",
    "MGR": "MUTED GREEN",
    "MUTED GREEN": "MUTED GREEN",
    "MLG": "MED LIGHT GREY",
    "MED LIGHT GREY": "MED LIGHT GREY",
    "MLN": "MED LIGHT NAVY",
    "MED LIGHT NAVY": "MED LIGHT NAVY",
    "MME": "MARSALA MELANGE",
    "MARSALA MELANGE": "MARSALA MELANGE",
    "MMM": "MACRO MESO MICRO",
    "MACRO MESO MICRO": "MACRO MESO MICRO",
    "MMX": "MIRAGE MATRIX",
    "MIRAGE MATRIX": "MIRAGE MATRIX",
    "MNB": "MIDNIGHT BLUE",
    "MIDNIGHT BLUE": "MIDNIGHT BLUE",
    "MON": "MOON",
    "MOON": "MOON",
    "MRN": "MAROON",
    "MAROON": "MAROON",
    "MSA": "MARSALA",
    "MARSALA": "MARSALA",
    "MSM": "MOLECULAR SPECTRUM",
    "MOLECULAR SPECTRUM": "MOLECULAR SPECTRUM",
    "MSS": "MELANGEANGE STRIPES",
    "MELANGEANGE STRIPES": "MELANGEANGE STRIPES",
    "MTE": "MOON TRACE",
    "MOON TRACE": "MOON TRACE",
    "MTG": "MINT GREEN",
    "MINT GREEN": "MINT GREEN",
    "MUE": "MAUVE",
    "MAUVE": "MAUVE",
    "MVO": "MIDNIGHT BLUE_VINTAGE INDIGO",
    "MIDNIGHT BLUE_VINTAGE INDIGO": "MIDNIGHT BLUE_VINTAGE INDIGO",
    "MYD": "MISTY DEPTHS",
    "MISTY DEPTHS": "MISTY DEPTHS",
    "MYE": "MYRTLE",
    "MYRTLE": "MYRTLE",
    "MYI": "MISTY INK",
    "MISTY INK": "MISTY INK",
    "NAM": "NORTH ATLANTIC MELANGE",
    "NORTH ATLANTIC MELANGE": "NORTH ATLANTIC MELANGE",
    "NBM": "NAVY B MELANGE",
    "NAVY B MELANGE": "NAVY B MELANGE",
    "NBO": "NAVY B OMPHALODES",
    "NAVY B OMPHALODES": "NAVY B OMPHALODES",
    "NBV": "NAVY B VINTAGE INDIGO",
    "NAVY B VINTAGE INDIGO": "NAVY B VINTAGE INDIGO",
    "NCN": "NEW CARBON",
    "NEW CARBON": "NEW CARBON",
    "NGN": "NEON GREEN",
    "NEON GREEN": "NEON GREEN",
    "NLB": "NAVY B_LT GREY B",
    "NAVY B_LT GREY B": "NAVY B_LT GREY B",
    "NPK": "NEON PINK",
    "NEON PINK": "NEON PINK",
    "NPY": "NAVY PEONY",
    "NAVY PEONY": "NAVY PEONY",
    "NSY": "NIGHT SKY",
    "NIGHT SKY": "NIGHT SKY",
    "NTC": "NORTH ATLANTIC",
    "NORTH ATLANTIC": "NORTH ATLANTIC",
    "NTM": "NIGHTMARE",
    "NIGHTMARE": "NIGHTMARE",
    "NVM": "NAVY MELANGE",
    "NAVY MELANGE": "NAVY MELANGE",
    "NVY": "NAVY",
    "NAVY": "NAVY",
    "NYB": "NAVY B",
    "NAVY B": "NAVY B",
    "NYH": "NAVY HUSH",
    "NAVY HUSH": "NAVY HUSH",
    "OBM": "OX BLOOD RED MELANGE",
    "OX BLOOD RED MELANGE": "OX BLOOD RED MELANGE",
    "OBR": "OX BLOOD RED",
    "OX BLOOD RED": "OX BLOOD RED",
    "OCT": "OVERCAST",
    "OVERCAST": "OVERCAST",
    "OFW": "OFF WHITE",
    "OFF WHITE": "OFF WHITE",
    "OGM": "OYSTER GRAY MELANGE",
    "OYSTER GRAY MELANGE": "OYSTER GRAY MELANGE",
    "OGY": "OYSTER GRAY",
    "OYSTER GRAY": "OYSTER GRAY",
    "OIN": "ONION",
    "ONION": "ONION",
    "OLD": "OMPHALODES",
    "OMPHALODES": "OMPHALODES",
    "OLM": "OMPHALODES MELANGE",
    "OMPHALODES MELANGE": "OMPHALODES MELANGE",
    "OLT": "OVERLAND TREK",
    "OVERLAND TREK": "OVERLAND TREK",
    "OME": "ONION MELANGE",
    "ONION MELANGE": "ONION MELANGE",
    "OML": "OAT MEAL",
    "OAT MEAL": "OAT MEAL",
    "OMS": "OMINOUS",
    "OMINOUS": "OMINOUS",
    "ONW": "OCEAN WAVES",
    "OCEAN WAVES": "OCEAN WAVES",
    "ORG": "ORANGE",
    "ORANGE": "ORANGE",
    "OSE": "OLIVE SMUDGE",
    "OLIVE SMUDGE": "OLIVE SMUDGE",
    "OVE": "OLIVE",
    "OLIVE": "OLIVE",
    "OVM": "OLIVE MELANGE",
    "OLIVE MELANGE": "OLIVE MELANGE",
    "PAE": "PINE APPLE",
    "PINE APPLE": "PINE APPLE",
    "PBM": "POOL BLUE MELANGE",
    "POOL BLUE MELANGE": "POOL BLUE MELANGE",
    "PCA": "PLUM CASPA",
    "PLUM CASPA": "PLUM CASPA",
    "PCH": "PEACH",
    "PEACH": "PEACH",
    "PCL": "PEACH CARAMEL",
    "PEACH CARAMEL": "PEACH CARAMEL",
    "PDB": "POWDER BLUE",
    "POWDER BLUE": "POWDER BLUE",
    "PDS": "PIXELATED STEPS",
    "PIXELATED STEPS": "PIXELATED STEPS",
    "PEM": "POWDER BLUE MELANGE",
    "POWDER BLUE MELANGE": "POWDER BLUE MELANGE",
    "PEP": "PURPLE PUDDLE",
    "PURPLE PUDDLE": "PURPLE PUDDLE",
    "PET": "PURPLE EFFECT",
    "PURPLE EFFECT": "PURPLE EFFECT",
    "PFT": "PINE FOREST",
    "PINE FOREST": "PINE FOREST",
    "PGI": "PIXELATED GRAFFITI",
    "PIXELATED GRAFFITI": "PIXELATED GRAFFITI",
    "PGM": "PINE GREEN MELANGE",
    "PINE GREEN MELANGE": "PINE GREEN MELANGE",
    "PGN": "PINE GREEN",
    "PINE GREEN": "PINE GREEN",
    "PHM": "PEACH MELANGE",
    "PEACH MELANGE": "PEACH MELANGE",
    "PHN": "PAINTED HORIZON",
    "PAINTED HORIZON": "PAINTED HORIZON",
    "PKN": "PUMPKIN",
    "PUMPKIN": "PUMPKIN",
    "PKP": "PINK PLUMAGE",
    "PINK PLUMAGE": "PINK PLUMAGE",
    "PLA": "PURPLE A",
    "PURPLE A": "PURPLE A",
    "PLB": "POOL BLUE",
    "POOL BLUE": "POOL BLUE",
    "PLM": "PURPLE A MELANGE",
    "PURPLE A MELANGE": "PURPLE A MELANGE",
    "PLP": "PARALLEL PULSE",
    "PARALLEL PULSE": "PARALLEL PULSE",
    "PLS": "PURPLE LINES",
    "PURPLE LINES": "PURPLE LINES",
    "PME": "PUMPKIN MELANGE",
    "PUMPKIN MELANGE": "PUMPKIN MELANGE",
    "PMT": "PURPLE MIST",
    "PURPLE MIST": "PURPLE MIST",
    "PNK": "PINK",
    "PINK": "PINK",
    "PPA": "PAPAYA",
    "PAPAYA": "PAPAYA",
    "PPE": "POTENT PURPLE",
    "POTENT PURPLE": "POTENT PURPLE",
    "PSA": "PISTA",
    "PISTA": "PISTA",
    "PSE": "PURPLE SMOKE",
    "PURPLE SMOKE": "PURPLE SMOKE",
    "PSL": "PINKY SWIRL",
    "PINKY SWIRL": "PINKY SWIRL",
    "PSS": "PURPLE SHARDS",
    "PURPLE SHARDS": "PURPLE SHARDS",
    "PSY": "PURPLE SPRAY",
    "PURPLE SPRAY": "PURPLE SPRAY",
    "PWD": "PLYWOOD",
    "PLYWOOD": "PLYWOOD",
    "RAB": "RAIN BLUE",
    "RAIN BLUE": "RAIN BLUE",
    "RBD": "RUBY RED",
    "RUBY RED": "RUBY RED",
    "RBM": "RASBERRY RADIENCE MIL",
    "RASBERRY RADIENCE MIL": "RASBERRY RADIENCE MIL",
    "RBN": "RUSTIC BROWN",
    "RUSTIC BROWN": "RUSTIC BROWN",
    "RDS": "RED SMOKE",
    "RED SMOKE": "RED SMOKE",
    "RDT": "RHOMBUS DRIFT",
    "RHOMBUS DRIFT": "RHOMBUS DRIFT",
    "RED": "RED",
    "RED": "RED",
    "REH": "RIDGE HUSH",
    "RIDGE HUSH": "RIDGE HUSH",
    "REM": "RED MELANGE",
    "RED MELANGE": "RED MELANGE",
    "RFT": "RAIN FOREST",
    "RAIN FOREST": "RAIN FOREST",
    "RGS": "RAIN GLASS",
    "RAIN GLASS": "RAIN GLASS",
    "RLK": "ROYAL PINK",
    "ROYAL PINK": "ROYAL PINK",
    "RLM": "ROYAL BLUE MELANGE",
    "ROYAL BLUE MELANGE": "ROYAL BLUE MELANGE",
    "RON": "RHODOENDRAN",
    "RHODOENDRAN": "RHODOENDRAN",
    "RPK": "ROSE PINK",
    "ROSE PINK": "ROSE PINK",
    "RRE": "RUM RAISIN MELANGE",
    "RUM RAISIN MELANGE": "RUM RAISIN MELANGE",
    "RRF": "ROSY REEF",
    "ROSY REEF": "ROSY REEF",
    "RRM": "RIPPLE REALM",
    "RIPPLE REALM": "RIPPLE REALM",
    "RRN": "RUM RAISIN",
    "RUM RAISIN": "RUM RAISIN",
    "RSE": "RED SMUDGE",
    "RED SMUDGE": "RED SMUDGE",
    "RSM": "RUST MELANGE",
    "RUST MELANGE": "RUST MELANGE",
    "RSR": "RASPBERRY RADIANCE",
    "RASPBERRY RADIANCE": "RASPBERRY RADIANCE",
    "RSS": "REGAL STREAKS",
    "REGAL STREAKS": "REGAL STREAKS",
    "RST": "RUST",
    "RUST": "RUST",
    "RVA": "RIVERA",
    "RIVERA": "RIVERA",
    "RVM": "RIVERA MELANGE",
    "RIVERA MELANGE": "RIVERA MELANGE",
    "RYB": "ROYAL BLUE",
    "ROYAL BLUE": "ROYAL BLUE",
    "SAD": "SAND",
    "SAND": "SAND",
    "SAN": "SEA GREEN",
    "SEA GREEN": "SEA GREEN",
    "SBE": "SMOKE BLUE",
    "SMOKE BLUE": "SMOKE BLUE",
    "SBK": "SILVER BRICK",
    "SILVER BRICK": "SILVER BRICK",
    "SBM": "SCAL BLUE MELANGE",
    "SCAL BLUE MELANGE": "SCAL BLUE MELANGE",
    "SCB": "SCAL BLUE",
    "SCAL BLUE": "SCAL BLUE",
    "SCT": "SHAPECRAFT",
    "SHAPECRAFT": "SHAPECRAFT",
    "SDE": "SAND DUNE",
    "SAND DUNE": "SAND DUNE",
    "SDL": "SHADOW LIME",
    "SHADOW LIME": "SHADOW LIME",
    "SDM": "SHADOW LIME B MELANGE",
    "SHADOW LIME B MELANGE": "SHADOW LIME B MELANGE",
    "SEG": "STONE GREY",
    "STONE GREY": "STONE GREY",
    "SES": "SHAPE SHIFTER",
    "SHAPE SHIFTER": "SHAPE SHIFTER",
    "SFR": "SUN FLOWER",
    "SUN FLOWER": "SUN FLOWER",
    "SGE": "SNARKEL GREEN MELANGE",
    "SNARKEL GREEN MELANGE": "SNARKEL GREEN MELANGE",
    "SGM": "STONE GREY MELANGE",
    "STONE GREY MELANGE": "STONE GREY MELANGE",
    "SGN": "SNARKEL GREEN",
    "SNARKEL GREEN": "SNARKEL GREEN",
    "SGY": "SILVER GREY",
    "SILVER GREY": "SILVER GREY",
    "SKM": "SKY BLUE MELANGE",
    "SKY BLUE MELANGE": "SKY BLUE MELANGE",
    "SKO": "SMOKEY OLIVE",
    "SMOKEY OLIVE": "SMOKEY OLIVE",
    "SKY": "SKY BLUE",
    "SKY BLUE": "SKY BLUE",
    "SLB": "SCHOOL BLUE",
    "SCHOOL BLUE": "SCHOOL BLUE",
    "SLC": "SEAMLESS CROSS",
    "SEAMLESS CROSS": "SEAMLESS CROSS",
    "SLD": "SEAMLESS DIAMOND",
    "SEAMLESS DIAMOND": "SEAMLESS DIAMOND",
    "SLE": "SALUTE",
    "SALUTE": "SALUTE",
    "SLG": "SEAMLESS GEOMETRICAL",
    "SEAMLESS GEOMETRICAL": "SEAMLESS GEOMETRICAL",
    "SLM": "SALUTE MELANGE",
    "SALUTE MELANGE": "SALUTE MELANGE",
    "SME": "SCHOOL BLUE MELANGE",
    "SCHOOL BLUE MELANGE": "SCHOOL BLUE MELANGE",
    "SNM": "SPACE NAVY MELANGE",
    "SPACE NAVY MELANGE": "SPACE NAVY MELANGE",
    "SNT": "STARRY NIGHT",
    "STARRY NIGHT": "STARRY NIGHT",
    "SOM": "STORM",
    "STORM": "STORM",
    "SPE": "STICK PUZZLE",
    "STICK PUZZLE": "STICK PUZZLE",
    "SPG": "SPRING GREEN",
    "SPRING GREEN": "SPRING GREEN",
    "SPK": "SILVER PINK",
    "SILVER PINK": "SILVER PINK",
    "SPM": "SPACE NAVY MEANGE",
    "SPACE NAVY MEANGE": "SPACE NAVY MEANGE",
    "SPN": "SPACE NAVY",
    "SPACE NAVY": "SPACE NAVY",
    "SRG": "SILVER GREY MELANGE",
    "SILVER GREY MELANGE": "SILVER GREY MELANGE",
    "SRT": "SKY ROCKET",
    "SKY ROCKET": "SKY ROCKET",
    "SSD": "SUMMER SAND",
    "SUMMER SAND": "SUMMER SAND",
    "SSE": "SHADED SPRUCE",
    "SHADED SPRUCE": "SHADED SPRUCE",
    "SSL": "SET SAIL",
    "SET SAIL": "SET SAIL",
    "SSN": "SHADOW SQUADRON",
    "SHADOW SQUADRON": "SHADOW SQUADRON",
    "SSR": "SOOTY SMEAR",
    "SOOTY SMEAR": "SOOTY SMEAR",
    "STE": "SKY TRACE",
    "SKY TRACE": "SKY TRACE",
    "SWP": "SHADOW PURPLE",
    "SHADOW PURPLE": "SHADOW PURPLE",
    "SWR": "STORMY WEATHER",
    "STORMY WEATHER": "STORMY WEATHER",
    "SWS": "SMOKE WAVES",
    "SMOKE WAVES": "SMOKE WAVES",
    "SYM": "SMOKEY OLIVE MELANGE",
    "SMOKEY OLIVE MELANGE": "SMOKEY OLIVE MELANGE",
    "SYN": "STARRY NIGHT",
    "STARRY NIGHT": "STARRY NIGHT",
    "SYP": "SMOKY PLATOON",
    "SMOKY PLATOON": "SMOKY PLATOON",
    "SYS": "SHADY SPUME",
    "SHADY SPUME": "SHADY SPUME",
    "TAM": "TARMAC MELANGE",
    "TARMAC MELANGE": "TARMAC MELANGE",
    "TAS": "TRINAGLES",
    "TRINAGLES": "TRINAGLES",
    "TCA": "TERRACOTTA",
    "TERRACOTTA": "TERRACOTTA",
    "TCM": "TERRACOTTA MELANGE",
    "TERRACOTTA MELANGE": "TERRACOTTA MELANGE",
    "TEL": "TEAL",
    "TEAL": "TEAL",
    "TEM": "TURBULENCE MELANGE",
    "TURBULENCE MELANGE": "TURBULENCE MELANGE",
    "TGE": "TEAL GRUNGE",
    "TEAL GRUNGE": "TEAL GRUNGE",
    "TGM": "TREKKING GREEN MELANGE",
    "TREKKING GREEN MELANGE": "TREKKING GREEN MELANGE",
    "TGN": "TANGERINE",
    "TANGERINE": "TANGERINE",
    "TIL": "TRI LOG",
    "TRI LOG": "TRI LOG",
    "TKM": "TURKISH MELANGE",
    "TURKISH MELANGE": "TURKISH MELANGE",
    "TLE": "TURBULENCE",
    "TURBULENCE": "TURBULENCE",
    "TLM": "TEAL MELANGE",
    "TEAL MELANGE": "TEAL MELANGE",
    "TLT": "TEAL TRACKS",
    "TEAL TRACKS": "TEAL TRACKS",
    "TMC": "TARMAC",
    "TARMAC": "TARMAC",
    "TME": "TROPOSPHERE MELANGE",
    "TROPOSPHERE MELANGE": "TROPOSPHERE MELANGE",
    "TNS": "TENNIS",
    "TENNIS": "TENNIS",
    "TPN": "TYPHOON",
    "TYPHOON": "TYPHOON",
    "TRN": "TREKKING GREEN",
    "TREKKING GREEN": "TREKKING GREEN",
    "TSM": "THUNDER STORM",
    "THUNDER STORM": "THUNDER STORM",
    "TSP": "TROPOSPHERE",
    "TROPOSPHERE": "TROPOSPHERE",
    "TTS": "TABLE TENNIS",
    "TABLE TENNIS": "TABLE TENNIS",
    "TTT": "TIC TAC TOE",
    "TIC TAC TOE": "TIC TAC TOE",
    "UGM": "ULTIMATE GREY MELANGE",
    "ULTIMATE GREY MELANGE": "ULTIMATE GREY MELANGE",
    "UGY": "ULTIMATE GREY",
    "ULTIMATE GREY": "ULTIMATE GREY",
    "VIM": "VINTAGE INDIGO MELANGE",
    "VINTAGE INDIGO MELANGE": "VINTAGE INDIGO MELANGE",
    "VIO": "VINTAGE INDIGO",
    "VINTAGE INDIGO": "VINTAGE INDIGO",
    "VNB": "VINTAGE INDIGO-NAVY B",
    "VINTAGE INDIGO-NAVY B": "VINTAGE INDIGO-NAVY B",
    "VTX": "VORTEX",
    "VORTEX": "VORTEX",
    "VWL": "VINTAGE WALL",
    "VINTAGE WALL": "VINTAGE WALL",
    "WAM": "WATER AQUA MELANGE",
    "WATER AQUA MELANGE": "WATER AQUA MELANGE",
    "WAQ": "WATER AQUA",
    "WATER AQUA": "WATER AQUA",
    "WCB": "WHITE_CAPRI BLUE",
    "WHITE_CAPRI BLUE": "WHITE_CAPRI BLUE",
    "WCR": "WATER COLOUR",
    "WATER COLOUR": "WATER COLOUR",
    "WDK": "WILD DUSK",
    "WILD DUSK": "WILD DUSK",
    "WDM": "WOODLAND GRAY MELANGE",
    "WOODLAND GRAY MELANGE": "WOODLAND GRAY MELANGE",
    "WFE": "WILD FIRE",
    "WILD FIRE": "WILD FIRE",
    "WFS": "WIND FORTRESS",
    "WIND FORTRESS": "WIND FORTRESS",
    "WGM": "WINTER GREEN MELANGE",
    "WINTER GREEN MELANGE": "WINTER GREEN MELANGE",
    "WGN": "WINTER GREEN",
    "WINTER GREEN": "WINTER GREEN",
    "WGW": "WAVEGLOW",
    "WAVEGLOW": "WAVEGLOW",
    "WHT": "WHITE",
    "WHITE": "WHITE",
    "WLG": "WOODLAND GRAY",
    "WOODLAND GRAY": "WOODLAND GRAY",
    "WPR": "WALL PAPER",
    "WALL PAPER": "WALL PAPER",
    "WWD": "WILD WIND",
    "WILD WIND": "WILD WIND",
    "WWH": "WATER WASH",
    "WATER WASH": "WATER WASH",
    "WWM": "WILD WIND MELANGE",
    "WILD WIND MELANGE": "WILD WIND MELANGE",
    "WYL": "WAX YELLOW",
    "WAX YELLOW": "WAX YELLOW",
    "YLM": "YELLOW MELANGE",
    "YELLOW MELANGE": "YELLOW MELANGE",
    "YLW": "YELLOW",
    "YELLOW": "YELLOW",
    "ZZH": "ZIGZAG ZENITH",
    "ZIGZAG ZENITH": "ZIGZAG ZENITH",
}


gsheet_color_map = {}

def load_dynamic_color_map():
    url = "https://docs.google.com/spreadsheets/d/1IjrPzNWndIzF0Cc4StaFmSNNFk2xW6icW47K8Th6F-U/export?format=xlsx"
    dynamic_map = {}
    try:
        import urllib.request
        import pandas as pd
        import io
        
        req = urllib.request.Request(url, headers={'User-Agent': 'Mozilla/5.0'})
        with urllib.request.urlopen(req, timeout=10) as response:
            xls = pd.ExcelFile(io.BytesIO(response.read()))
            
            # Read 'COLOUR CODE SHEET'
            if 'COLOUR CODE SHEET' in xls.sheet_names:
                df = pd.read_excel(xls, sheet_name='COLOUR CODE SHEET')
                for _, row in df.iterrows():
                    color = str(row.get('COLOR', '')).strip().upper()
                    code = str(row.get('CODE', '')).strip().upper()
                    if color and code and color != "NAN" and code != "NAN":
                        dynamic_map[code] = color
                        dynamic_map[color] = color
            
            # Also read 'EAN' sheet for any extra color mappings
            if 'EAN' in xls.sheet_names:
                df = pd.read_excel(xls, sheet_name='EAN')
                for _, row in df.iterrows():
                    color = str(row.get('COLOR', '')).strip().upper()
                    code = str(row.get('CODE', '')).strip().upper()
                    if color and code and color != "NAN" and code != "NAN":
                        dynamic_map[code] = color
                        dynamic_map[color] = color
                        
        print(f"Loaded {len(dynamic_map)} color mappings dynamically from Google Sheets.")
    except Exception as e:
        print(f"Warning: Could not load dynamic color map: {e}")
        
    return dynamic_map


def normalize_color(x):
    if not x:
        return ""
    c = str(x).strip().upper().replace("GREY", "GRAY").replace("LIGTH", "LIGHT").replace("CHILLI", "CHILI").replace("BLUESTONE", "BLUE STONE").replace("FUSIA", "FUSHIA").replace("_", " ").replace("-", " ")
    
    if c.startswith("LT "):
        c = "LIGHT " + c[3:]
    c = c.replace(" LT ", " LIGHT ")
    
    if c.startswith("DK "):
        c = "DARK " + c[3:]
    c = c.replace(" DK ", " DARK ")

    descriptors = {"PRO", "NEO", "PLUS", "PREMIUM", "LITE", "MAX", "ULTRA", "SPORT", "ACTIVE", "EDITION", "SERIES", "FIT", "CLASSIC", "FLEX", "PRIME", "STUDIO", "COLLECTION", "LINE", "AIR", "TECH", "DRY"}
    
    abbrevs = {
        "LT": "LIGHT",
        "DK": "DARK",
        "BB": "BLUE BERRY",
        "LIGTH": "LIGHT"
    }
    
    words = []
    for w in c.split():
        if w in descriptors:
            continue
        words.append(abbrevs.get(w, w))
    c = " ".join(words)
            
    res = gsheet_color_map.get(c)
    if not res:
        res = color_map.get(c, c)
    res_str = str(res).strip().upper().replace("GREY", "GRAY").replace("_", " ").replace("-", " ")
    
    if res_str.startswith("LT "):
        res_str = "LIGHT " + res_str[3:]
    res_str = res_str.replace(" LT ", " LIGHT ")
    
    if res_str.startswith("DK "):
        res_str = "DARK " + res_str[3:]
    res_str = res_str.replace(" DK ", " DARK ")

    res_words = []
    for w in res_str.split():
        if w in descriptors:
            continue
        res_words.append(abbrevs.get(w, w))
    res_str = " ".join(res_words)

    for var_suffix in [" A", " B", " C", " D"]:
        if res_str.endswith(var_suffix):
            res_str = res_str[:-len(var_suffix)].strip()
            
    return res_str


def normalize_category_value(x):
    if not x:
        return ""
    s = str(x).strip().upper().replace("'", "").replace("’", "")
    if "WOMEN" in s:
        return "WOMENS"
    if "MEN" in s:
        return "MENS"
    if "BOY" in s:
        return "BOYS"
    if "GIRL" in s:
        return "GIRLS"
    if "KID" in s:
        return "KIDS"
    return s


def normalize_number(x):
    if x is None or x == "":
        return None
    s = str(x).replace(",", "").replace("₹", "").strip()
    if "pair" in s.lower():
        return 1.0
    m = re.search(r"[\d,]+\.?\d*", s)
    if m:
        try:
            return round(float(m.group().replace(",", "")), 2)
        except ValueError:
            pass
    return None


def find_col(df, *candidates):
    """Find a column in df whose name loosely matches one of the candidates."""
    cols_upper = {str(c).upper().strip(): c for c in df.columns}
    for cand in candidates:
        cand_u = cand.upper().strip()
        if cand_u in cols_upper:
            return cols_upper[cand_u]
    for cand in candidates:
        cand_u = cand.upper().strip()
        for cu, orig in cols_upper.items():
            if cand_u in cu:
                # Exclude columns containing dates/locations/totals for MRP
                if cand_u == "MRP" and any(x in cu for x in ["DATE", "LOCATION", "ACTIVE", "TOTAL", "BOX", "BUNDLE", "PACK"]):
                    continue
                return orig
    # Fallback to match anything if no clean match found
    for cand in candidates:
        cand_u = cand.upper().strip()
        for cu, orig in cols_upper.items():
            if cand_u in cu:
                return orig
    return None


def clean_style_for_gsheet(style):
    if not style:
        return ""
    s = str(style).strip().upper()
    if s.startswith("OR"):
        s = s[1:]
    elif s.startswith("SOR"):
        s = s[2:]
    return s
def parse_fit_from_text(text):
    import pandas as pd
    if not text or pd.isna(text):
        return None
    t = str(text).upper()
    if "REGULAR" in t:
        return "Regular Fit"
    if "SLIM" in t:
        return "Slim Fit"
    if "OVERSIZED" in t or "OVERSIZE" in t:
        return "Oversized Fit"
    return None



def detect_gender_from_sku(sku):
    s = str(sku).strip().upper()
    if not s:
        return None
    first_char = s[0]
    if first_char == "W":
        return "WOMENS"
    if first_char == "M":
        return "MENS"
    if first_char == "B":
        return "BOYS"
    if first_char == "G":
        return "GIRLS"
    if first_char == "K":
        return "KIDS"
    # Fallback to check size
    try:
        _, _, size = extract_sku_details(s)
        if size and "Y" in size:
            return "KIDS"
    except Exception:
        pass
    return None


def gender_matches(row_gender, sku_gender):
    if not sku_gender or not row_gender:
        return True
    rg = str(row_gender).strip().upper()
    sg = sku_gender.strip().upper()
    
    # Handle typos in Google Sheets
    if "WOMN" in rg or "WMN" in rg or "WOMEN" in rg:
        rg = "WOMENS"
    if "WOMN" in sg or "WMN" in sg or "WOMEN" in sg:
        sg = "WOMENS"
        
    if "MEN" in rg and rg != "WOMENS":
        rg = "MENS"
    if "MEN" in sg and sg != "WOMENS":
        sg = "MENS"

    if rg in ["MENS", "WOMENS"] and sg in ["MENS", "WOMENS"]:
        return rg == sg

    if sg in ["KIDS", "BOYS", "GIRLS"] or any(k in sg for k in ["KID", "BOY", "GIRL"]):
        return any(k in rg for k in ["KID", "BOY", "GIRL"])
    return sg == rg


def match_style_code(pdf_style_base, gs_style_base, tag_type="Standard Garment / Dress Tags"):
    p_s = str(pdf_style_base).strip().upper()
    g_s = str(gs_style_base).strip().upper()
    
    if p_s == g_s:
        return True
        
    p_prefix = "".join([c for c in p_s if c.isalpha()])
    g_prefix = "".join([c for c in g_s if c.isalpha()])
    
    p_clean = clean_prefix(p_prefix)
    g_clean = clean_prefix(g_prefix)
    if p_clean and g_clean and p_clean != g_clean:
        return False
        
    if tag_type == "D2C Dress tag file":
        if p_prefix and g_prefix and p_prefix != g_prefix:
            if not ((p_prefix == "S" and g_prefix == "WS") or (p_prefix == "WS" and g_prefix == "S")):
                return False
                
    p_digits = "".join([c for c in p_s if c.isdigit()])
    g_digits = "".join([c for c in g_s if c.isdigit()])
    
    return p_digits == g_digits


def match_batch_code(pdf_batch, gs_batch):
    p_b = str(pdf_batch).strip().upper()
    g_b = str(gs_batch).strip().upper()
    
    p_digits = "".join([c for c in p_b if c.isdigit()])
    g_digits = "".join([c for c in g_b if c.isdigit()])
    
    p_alpha = "".join([c for c in p_b if c.isalpha()])
    g_alpha = "".join([c for c in g_b if c.isalpha()])
    
    def get_val(dig):
        if not dig:
            return 1
        try:
            val = int(dig)
            return 1 if val == 0 else val
        except ValueError:
            return -1
            
    p_val = get_val(p_digits)
    g_val = get_val(g_digits)
    
    if p_val != g_val:
        return False
        
    if (not p_alpha or not g_alpha) and (p_alpha or g_alpha):
        non_empty = p_alpha if p_alpha else g_alpha
        if non_empty not in ["DT", "SB"]:
            return False
    elif p_alpha != g_alpha:
        return False
        
    return True


def append_sku_batch_to_style(pdf_style, pdf_sku):
    style_clean = str(pdf_style).strip().upper() if pdf_style else ""
    sku_clean = str(pdf_sku).strip().upper() if pdf_sku else ""

    if "/" not in style_clean and sku_clean:
        try:
            n = len(sku_clean)
            rules = {
                11: (2, 4, 0),
                12: (2, 4, 0),
                13: (2, 5, 0),
                14: (2, 4, 2),
                15: (2, 4, 3),
                16: (2, 5, 3),
                17: (2, 8, 0),
                18: (3, 6, 3),
            }
            if n in rules:
                _, _, end_remove = rules[n]
                if end_remove > 0:
                    suffix = sku_clean[-end_remove:]
                    has_alpha = any(c.isalpha() for c in suffix)
                    if has_alpha:
                        import re
                        match = re.match(r"^([A-Z]+)(0*)([0-9]+)$", suffix)
                        if match:
                            batch_val = match.group(1) + match.group(3)
                        else:
                            batch_val = suffix
                    else:
                        suffix_digits = "".join([c for c in suffix if c.isdigit()])
                        if suffix_digits:
                            batch_val = str(int(suffix_digits))
                        else:
                            batch_val = suffix
                    style_clean = f"{style_clean}/{batch_val}"
        except Exception:
            pass
    return style_clean


def find_row_by_style_and_batch(df, pdf_style, pdf_sku, tag_type="Standard Garment / Dress Tags"):
    style_clean = str(pdf_style).strip().upper()
    
    # Extract parts
    p_parts = style_clean.split("/")
    p_style_base = p_parts[0].strip()
    p_batch = p_parts[1].strip() if len(p_parts) > 1 else ""
    
    # Check if SKU has suffix
    if not p_batch and pdf_sku:
        res = extract_sku_details_with_batch(pdf_sku)
        if res and len(res) == 4:
            p_batch = res[3]
                        
    sku_gender = detect_gender_from_sku(pdf_sku)
    
    gender_col = None
    for c in df.columns:
        if str(c).upper().strip() in ["GENDER", "590"]:
            gender_col = c
            break

    # First attempt: match with gender and exact prefix matching
    for _, row in df.iterrows():
        gs_style = str(row.get("STYLE NO", "")).strip().upper()
        gs_batch = str(row.get("BATCH", "")).strip().upper()
        if gs_style and gs_style != "NAN":
            if match_style_code(p_style_base, gs_style, tag_type):
                if gender_col is None or gender_matches(row.get(gender_col), sku_gender):
                    if match_batch_code(p_batch, gs_batch):
                        return row
                        
    # Fallback to match style and empty batch
    for _, row in df.iterrows():
        gs_style = str(row.get("STYLE NO", "")).strip().upper()
        gs_batch = str(row.get("BATCH", "")).strip().upper()
        if gs_style and gs_style != "NAN":
            if match_style_code(p_style_base, gs_style, tag_type):
                if gender_col is None or gender_matches(row.get(gender_col), sku_gender):
                    if not gs_batch or gs_batch == "NAN" or gs_batch == "0":
                        return row
                            
    return None


def clean_prefix(prefix):
    p = str(prefix).strip().upper()
    two_letter = ("MT", "WT", "MS", "WS", "MV", "WV", "MI", "WI", "MJ", "WJ", "WP", "MP", "WB", "MB", "BT", "GP", "KD")
    for tl in two_letter:
        if p.startswith(tl):
            p = p[len(tl):]
            break
    category_letters = {"O", "S", "P", "T", "M", "W", "K", "B", "G", "I", "J", "V", "D"}
    while len(p) > 0 and p[0] in category_letters:
        p = p[1:]
    return p


def get_updated_fit(pdf_style, pdf_sku, gsheet_dfs, tag_type="Standard Garment / Dress Tags"):
    if not gsheet_dfs:
        return None
    for sheet_name in ["DT FINAL MRP", "New MRP 26-27"]:
        df = gsheet_dfs.get(sheet_name)
        if df is not None:
            row = find_row_by_style_and_batch(df, pdf_style, pdf_sku, tag_type)
            if row is not None:
                fit = row.get("FIT")
                if pd.notna(fit):
                    return str(fit).strip()
    return None


def get_updated_mrp(pdf_style, pdf_sku, gsheet_dfs, tag_type="Standard Garment / Dress Tags"):
    if not gsheet_dfs:
        return None

    for sheet_name in ["DT FINAL MRP", "New MRP 26-27"]:
        df = gsheet_dfs.get(sheet_name)
        if df is not None:
            row = find_row_by_style_and_batch(df, pdf_style, pdf_sku, tag_type)
            if row is not None:
                mrp_val = row.get("MRP")
                if pd.notna(mrp_val):
                    return mrp_val
    return None


def get_updated_total_mrp(pdf_style, pdf_sku, gsheet_dfs, tag_type="Standard Garment / Dress Tags"):
    if not gsheet_dfs:
        return None

    for sheet_name in ["DT FINAL MRP", "New MRP 26-27"]:
        df = gsheet_dfs.get(sheet_name)
        if df is not None:
            row = find_row_by_style_and_batch(df, pdf_style, pdf_sku, tag_type)
            if row is not None:
                box_mrp = row.get("MRP.1")
                if pd.notna(box_mrp) and str(box_mrp).strip() and str(box_mrp).strip().upper() != "NAN":
                    try:
                        return float(box_mrp)
                    except (ValueError, TypeError):
                        pass
                unit_mrp = row.get("MRP")
                pcs = row.get("PCS PER BOXES")
                if pd.notna(unit_mrp) and pd.notna(pcs):
                    try:
                        return float(unit_mrp) * float(pcs)
                    except (ValueError, TypeError):
                        pass
    return None


def get_updated_lot_no(pdf_style, pdf_sku, gsheet_dfs, tag_type="Standard Garment / Dress Tags"):
    if not gsheet_dfs:
        return None

    for sheet_name in ["DT FINAL MRP", "New MRP 26-27"]:
        df = gsheet_dfs.get(sheet_name)
        if df is not None:
            row = find_row_by_style_and_batch(df, pdf_style, pdf_sku, tag_type)
            if row is not None:
                col_name = None
                for c in df.columns:
                    if str(c).strip() == ",":
                        col_name = c
                        break
                if col_name:
                    val = row.get(col_name)
                    if pd.notna(val) and str(val).strip():
                        return str(val).strip()
                
                gs_s = row.get("STYLE NO")
                gs_b = row.get("BATCH")
                b_suffix = f"/{gs_b}" if pd.notna(gs_b) and str(gs_b).strip() and str(gs_b).strip().upper() != "NAN" else ""
                return f"{gs_s}{b_suffix}"

    return None


def get_updated_description(pdf_style, pdf_sku, gsheet_dfs, tag_type="Standard Garment / Dress Tags"):
    if not gsheet_dfs:
        return None

    for sheet_name in ["DT FINAL MRP", "New MRP 26-27"]:
        df = gsheet_dfs.get(sheet_name)
        if df is not None:
            row = find_row_by_style_and_batch(df, pdf_style, pdf_sku, tag_type)
            if row is not None:
                desc = row.get("DESCRIPTION")
                if pd.notna(desc):
                    return str(desc).strip()
    return None


def get_updated_category(pdf_style, pdf_sku, gsheet_dfs, tag_type="Standard Garment / Dress Tags"):
    if not gsheet_dfs:
        return None

    for sheet_name in ["DT FINAL MRP", "New MRP 26-27"]:
        df = gsheet_dfs.get(sheet_name)
        if df is not None:
            row = find_row_by_style_and_batch(df, pdf_style, pdf_sku, tag_type)
            if row is not None:
                gender_col = None
                for c in df.columns:
                    if str(c).upper().strip() in ["GENDER", "590"]:
                        gender_col = c
                        break
                if gender_col:
                    g_val = row.get(gender_col)
                    if pd.notna(g_val):
                        g_str = str(g_val).strip().upper()
                        if "WOMEN" in g_str or "WOMN" in g_str or "WMN" in g_str:
                            return "Women's"
                        elif "MEN" in g_str:
                            return "Men's"
                        elif "BOY" in g_str:
                            return "Boy's"
                        elif "GIRL" in g_str:
                            return "Girl's"
                        elif "KID" in g_str:
                            return "Kid's"
    return None


def extract_sku_details_with_batch(sku_str):
    sku = str(sku_str).strip().upper()
    n = len(sku)

    size_keywords = ["08Y", "10Y", "12Y", "14Y", "02Y", "04Y", "06Y", "2Y", "4Y", "6Y", "8Y", "XSML", "SML", "MED", "LAR", "XLR", "2XLR", "3XLR", "4XLR", "5XLR", "XXL", "XXXL", "XXXXL", "2XL", "3XL", "4XL", "5XL", "XS", "S", "M", "L", "XL"]
    
    found_size = None
    size_pos = -1
    for sz_kw in size_keywords:
        pos = sku.find(sz_kw)
        if pos != -1:
            if pos >= 5:
                found_size = sz_kw
                size_pos = pos
                break
                
    if found_size:
        left = sku[:size_pos]
        batch = sku[size_pos + len(found_size):]
        
        color = left[-3:]
        rest = left[:-3]
        
        if rest.startswith(("MT", "WT", "MS", "WS", "MV", "WV", "MI", "WI", "MJ", "WJ", "WP", "MP", "WB", "MB", "BT", "GP", "KD")):
            style = rest[2:]
        elif rest.startswith(("M", "W", "K", "B", "G")):
            style = rest[1:]
        else:
            style = rest
            
        return style, color, found_size, batch

    # Fallback to legacy length rules
    if sku.endswith(("2PK", "3PK")):
        batch = sku[-6:]
        sku_without_batch = sku[:-6]
    else:
        m = re.search(r"\d+$", sku)
        if m:
            batch = m.group()
            sku_without_batch = sku[:-len(batch)]
        else:
            batch = ""
            sku_without_batch = sku

    found_size = None
    for sz_kw in size_keywords:
        if sku_without_batch.endswith(sz_kw):
            found_size = sz_kw
            break

    if found_size:
        size_idx = len(sku_without_batch) - len(found_size)
        left = sku_without_batch[:size_idx]
        if len(left) >= 5:
            color = left[-3:]
            rest = left[:-3]
            if rest.startswith(("MT", "WT", "MS", "WS", "MV", "WV", "MI", "WI", "MJ", "WJ", "WP", "MP", "WB", "MB", "BT", "GP", "KD")):
                style = rest[2:]
            elif rest.startswith(("M", "W", "K", "B", "G")):
                style = rest[1:]
            else:
                style = rest
            return style, color, found_size, batch

    rules = {
        11: (2, 4, 0),
        12: (2, 4, 0),
        13: (2, 5, 0),
        14: (2, 4, 2),
        15: (2, 4, 3),
        16: (2, 5, 3),
        17: (2, 8, 0),
        18: (2, 4, 6) if sku.endswith(("2PK", "3PK")) else (3, 6, 3),
    }

    if n not in rules:
        return None, None, None, ""

    if n == 12:
        body = sku[2:]
        if len(body) >= 5 and body[:2].isalpha() and body[2:5].isdigit():
            style_len = 5
        else:
            style_len = 4
        style = body[:style_len]
        if len(style) >= 3 and style[1:3] == "OR":
            style = style[1:]
        color = body[style_len:style_len+3]
        size = body[style_len+3:]
        return style, color, size, ""

    if n == 15:
        apparel_sizes = {"MED", "LAR", "XLR", "2XL", "3XL", "4XL", "5XL", "SML"}
        if sku[-3:] in apparel_sizes:
            style = sku[2:-6]
            if len(style) >= 3 and style[1:3] == "OR":
                style = style[1:]
            color = sku[-6:-3]
            size = sku[-3:]
            return style, color, size, ""
        elif sku[-6:-3] in apparel_sizes:
            style = sku[2:-9]
            if len(style) >= 3 and style[1:3] == "OR":
                style = style[1:]
            color = sku[-9:-6]
            size = sku[-6:-3]
            return style, color, size, ""

    start_remove, style_len, end_remove = rules[n]
    body = sku[start_remove:]
    if end_remove:
        body = body[:-end_remove]

    style = body[:style_len]
    if len(style) >= 3 and style[1:3] == "OR":
        style = style[1:]
    color = body[style_len:style_len+3]
    size = body[-3:]

    return style, color, size, ""


def extract_sku_details(sku_str):
    res = extract_sku_details_with_batch(sku_str)
    if res:
        return res[0], res[1], res[2]
    return None, None, None


def parse_product_name_info(prod_name_str):
    if not prod_name_str:
        return "", "", "", None
    parts = str(prod_name_str).strip().split()
    lot_no = parts[0] if parts else ""
    base_style = lot_no.split("/")[0] if "/" in lot_no else lot_no

    pcs_match = re.search(r"(\d+)\s*(PCS|NOS|PACK)", str(prod_name_str), re.IGNORECASE)
    if pcs_match:
        pcs_qty = int(pcs_match.group(1))
        matched_str = pcs_match.group(0)
    else:
        sizes_pattern = r"\b(\d+)\s+(XSML|SML|MED|LAR|XLR|2XLR|3XLR|4XLR|5XLR|XXL|XXXL|2XL|3XL|4XL|5XL|XS|S|M|L|XL|S/36|M/38|L/40|XL/42|2XL/44|3XL/46)\s*$"
        m = re.search(sizes_pattern, str(prod_name_str).strip(), re.IGNORECASE)
        if m:
            pcs_qty = int(m.group(1))
            matched_str = m.group(1)
        else:
            pcs_qty = None
            matched_str = None

    desc = " ".join(parts[1:])
    if matched_str:
        desc = desc.split(matched_str)[0].strip()
    desc = re.sub(r"\s+(ASSORTED|BLACK|NAVY|GREY|GRAY|WHITE)$", "", desc, flags=re.IGNORECASE).strip()

    return lot_no, base_style, desc, pcs_qty


def extract_style_and_size_from_sku(sku_str):
    style, color, size = extract_sku_details(sku_str)
    return style, size


PRODUCT_GROUPS = [
    {"PANT", "JOGGER", "LOWER", "TIGHT", "LEGGING", "CAPRI"},
    {"SHORT"},
    {"SHIRT", "TEE", "POLO", "CREWNECK", "JACKET", "HOODIE", "SWEATSHIRT", "BRA"},
    {"SOCKS"},
    {"SHOE"},
    {"BOXER", "INNERWEAR", "BRIEF"}
]

CLEAN_PRODUCT_NOUNS = {
    "PANT", "JOGGER", "LOWER", "TIGHT", "LEGGING", "CAPRI",
    "SHORT",
    "SHIRT", "TEE", "POLO", "CREWNECK", "JACKET", "HOODIE", "SWEATSHIRT", "BRA",
    "SOCKS", "SHOE", "BOXER", "INNERWEAR", "BRIEF"
}

def check_conflicting_product_type(pdf_words, excel_words):
    pdf_groups = set()
    excel_groups = set()
    for idx, grp in enumerate(PRODUCT_GROUPS):
        if any(w in grp for w in pdf_words):
            pdf_groups.add(idx)
        if any(w in grp for w in excel_words):
            excel_groups.add(idx)
            
    if pdf_groups and excel_groups and not pdf_groups.intersection(excel_groups):
        return True
    return False


def compare(pdf_df: pd.DataFrame, excel_df: pd.DataFrame, gsheet_dfs: dict, tag_type: str = "D2C Dress tag file") -> pd.DataFrame:
    global gsheet_color_map
    if not gsheet_color_map:
        gsheet_color_map = load_dynamic_color_map()
    desc_col = find_col(excel_df, "PRODUCT NAME", "DESCRIPTION", "PRODUCT")
    lot_col = find_col(excel_df, "LOT NO", "LOT", "STYLE", "STYLE CODE")
    barcode_col = find_col(excel_df, "BARCODE", "BAR CODE", "EAN", "GTIN")
    mrp_col = find_col(excel_df, "MRP")
    total_mrp_col = find_col(excel_df, "TOTAL MRP")
    size_col = find_col(excel_df, "SIZE")
    color_col = find_col(excel_df, "COLOUR", "COLOR")
    qty_col = find_col(excel_df, "PACK QTY", "NET QTY", "QTY", "TAG QTY")
    category_col = find_col(excel_df, "CATEGORY", "GENDER")

    sku_col = find_col(
        excel_df,
        "SKU CODE", "SKU NUMBER", "SKU NO", "SKU",
        "ITEM CODE", "ITEM_CODE", "ITEM NO", "ITEM NUMBER", "ITEM",
        "PRODUCT CODE", "MATERIAL NO", "MATERIAL", "ARTICLE NO", "ARTICLE", "CODE"
    )
    if sku_col is None and barcode_col is not None:
        sku_col = barcode_col
    if sku_col is None:
        sku_col = find_col(excel_df, "STYLE NO", "STYLE CODE", "STYLE")

    if sku_col is None:
        cols_preview = ", ".join(excel_df.columns[:10])
        raise ValueError(f"Could not find an SKU or Identifier column in the Excel sheet. Available columns: [{cols_preview}]")

    excel_idx_sku = {normalize_sku(row[sku_col]): row for _, row in excel_df.iterrows()}
    excel_idx_barcode = {}
    if barcode_col:
        for _, row in excel_df.iterrows():
            bar_val = row.get(barcode_col)
            if pd.notna(bar_val) and str(bar_val).strip():
                excel_idx_barcode[str(bar_val).strip()] = row

    if tag_type == "B2B Bundle Sticker tag file":
        field_map = [
            ("Color", color_col, normalize_color),
            ("Lot No (Google Sheet)", None, normalize_lot),
            ("Lot No (GS1 Master)", lot_col, normalize_lot),
            ("Qty", qty_col, normalize_number),
            ("Total MRP", None, normalize_number),
            ("SKU", sku_col, normalize_sku),
            ("EAN", barcode_col, normalize_text),
            ("Size", size_col, normalize_size),
        ]
    elif tag_type == "B2B Box Sticker tag file":
        field_map = [
            ("Description", desc_col, normalize_text),
            ("Lot No (Google Sheet)", None, normalize_lot),
            ("Lot No (GS1 Master)", lot_col, normalize_lot),
            ("Qty", qty_col, normalize_number),
            ("Total MRP", None, normalize_number),
            ("SKU", sku_col, normalize_sku),
            ("EAN", barcode_col, normalize_text),
            ("Size", size_col, normalize_size),
        ]
    else:
        field_map = [
            ("Description", desc_col, normalize_text),
            ("Fit", None, normalize_text),
            ("Category", category_col, normalize_category_value),
            ("MRP", None, normalize_number),
            ("SKU", sku_col, normalize_sku),
            ("EAN", barcode_col, normalize_text),
            ("Size", size_col, normalize_size),
            ("Color", color_col, normalize_color),
            ("Qty", qty_col, normalize_number),
            ("Lot No (Google Sheet)", None, normalize_lot),
            ("Lot No (GS1 Master)", lot_col, normalize_lot),
        ]

    report_rows = []
    matched_excel_skus = set()

    for _, tag in pdf_df.iterrows():
        pdf_sku_norm = normalize_sku(tag["SKU"])
        pdf_barcode = str(tag.get("EAN") or tag.get("Barcode") or "").strip()
        
        # 1. Lookup by SKU first
        excel_row = excel_idx_sku.get(pdf_sku_norm)
        
        # 2. Fallback to lookup by Barcode/GTIN if SKU is not found
        if excel_row is None and pdf_barcode:
            excel_row = excel_idx_barcode.get(pdf_barcode)

        is_simulated = False
        if excel_row is None:
            style, color_code, size_code = extract_sku_details(pdf_sku_norm)
            if style:
                is_simulated = True
                mapped_c = gsheet_color_map.get(color_code) or color_map.get(color_code, color_code)
                excel_row = {
                    "SKU": pdf_sku_norm,
                    "Size": size_code,
                    "Color": mapped_c,
                    "Style": style
                }
                if sku_col:
                    excel_row[sku_col] = pdf_sku_norm
                if barcode_col:
                    excel_row[barcode_col] = None
                if size_col:
                    excel_row[size_col] = size_code
                if color_col:
                    excel_row[color_col] = mapped_c
                if desc_col:
                    excel_row[desc_col] = None
                if lot_col:
                    excel_row[lot_col] = style
            else:
                report_rows.append({
                    "SKU": tag["SKU"],
                    "Field": "SKU",
                    "PDF Value": tag["SKU"],
                    "Excel Value": None,
                    "Status": "❌ Not found in Excel",
                })
                continue

        matched_excel_skus.add(pdf_sku_norm)

        prod_name_val = excel_row.get(desc_col) if desc_col else None
        lot_info, base_style_info, desc_info, pack_qty_info = parse_product_name_info(prod_name_val)

        for field_name, excel_col, norm_fn in field_map:
            if excel_col is None and field_name.endswith("(GS1 Master)"):
                continue

            pdf_val = tag.get(field_name)
            excel_val = None

            if field_name == "Description":
                pdf_val = tag.get("Description") or tag.get("Product")
                if not pdf_val:
                    pdf_val = desc_info
                excel_val = desc_info if desc_info else (excel_row.get(excel_col) if excel_col else None)
            elif field_name == "Lot No (Google Sheet)":
                pdf_val = tag.get("Lot No") or tag.get("Style")
                g_lot = get_updated_lot_no(tag.get("Style") or base_style_info, tag.get("SKU"), gsheet_dfs, tag_type=tag_type)
                excel_val = g_lot if g_lot and pd.notna(g_lot) and str(g_lot).strip() != "" and str(g_lot).strip().upper() != "NAN" else (tag.get("Style") or base_style_info)
            elif field_name == "Lot No (GS1 Master)":
                pdf_val = tag.get("Lot No") or tag.get("Style")
                db_lot = lot_info if lot_info else (excel_row.get(excel_col) if excel_col else None)
                excel_val = db_lot if db_lot and pd.notna(db_lot) and str(db_lot).strip() != "" and str(db_lot).strip().upper() != "NAN" else (tag.get("Style") or base_style_info)
            elif field_name == "Qty":
                pdf_val = tag.get("Net Quantity") or tag.get("Qty")
                excel_val = pack_qty_info if pack_qty_info else (excel_row.get(excel_col) if excel_col else 1.0)
            elif field_name == "MRP":
                pdf_val = tag.get("MRP")
                g_mrp = get_updated_mrp(tag.get("Style") or base_style_info, tag.get("SKU"), gsheet_dfs, tag_type=tag_type)
                excel_val = g_mrp if g_mrp else (excel_row.get(mrp_col) if mrp_col else None)
            elif field_name == "Total MRP":
                pdf_val = tag.get("Total MRP")
                if pdf_val is None:
                    continue
                # 1. Fetch authoritative Total MRP directly from Google Sheet (MRP.1 / PCS PER BOXES)
                g_total_mrp = get_updated_total_mrp(tag.get("Style") or base_style_info, tag.get("SKU"), gsheet_dfs, tag_type=tag_type)
                if g_total_mrp is not None:
                    excel_val = g_total_mrp
                else:
                    excel_val = excel_row.get(total_mrp_col) if (total_mrp_col and pd.notna(excel_row.get(total_mrp_col)) and str(excel_row.get(total_mrp_col)).strip() != "" and str(excel_row.get(total_mrp_col)).strip().upper() != "NAN") else None
                    if excel_val is None:
                        single_mrp = get_updated_mrp(tag.get("Style") or base_style_info, tag.get("SKU"), gsheet_dfs, tag_type=tag_type)
                        if not single_mrp:
                            single_mrp = excel_row.get(mrp_col) if mrp_col else None
                        if tag_type == "B2B Box Sticker tag file":
                            p_qty = norm_fn(tag.get("Net Quantity")) or pack_qty_info or 8.0
                        elif tag_type == "B2B Bundle Sticker tag file":
                            p_qty = pack_qty_info or norm_fn(excel_row.get(qty_col)) or 5.0
                        else:
                            p_qty = norm_fn(tag.get("Qty") or tag.get("Net Quantity")) or 1.0
                        if single_mrp and p_qty:
                            try:
                                excel_val = float(single_mrp) * float(p_qty)
                            except (ValueError, TypeError):
                                excel_val = None
                        else:
                            excel_val = None
            elif field_name == "Fit":
                pdf_val = tag.get("Fit")
                g_fit = get_updated_fit(tag.get("Style") or base_style_info, tag.get("SKU"), gsheet_dfs, tag_type=tag_type)
                if g_fit:
                    excel_val = g_fit
                else:
                    prod_name = excel_row.get(desc_col) if desc_col else None
                    prod_desc = excel_row.get("Product Description")
                    excel_val = parse_fit_from_text(prod_name) or parse_fit_from_text(prod_desc)
            elif field_name == "Category":
                pdf_val = tag.get("Category")
                if excel_col:
                    excel_val = excel_row.get(excel_col)
                else:
                    g_cat = get_updated_category(tag.get("Style") or base_style_info, tag.get("SKU"), gsheet_dfs, tag_type=tag_type)
                    if g_cat:
                        excel_val = g_cat
                    else:
                        prod_text = (str(excel_row.get(desc_col) or "") + " " + str(excel_row.get("Product Description") or "")).upper()
                        if "WOMEN" in prod_text:
                            excel_val = "Women's"
                        elif "MEN" in prod_text:
                            excel_val = "Men's"
                        elif "BOY" in prod_text:
                            excel_val = "Boy's"
                        elif "GIRL" in prod_text:
                            excel_val = "Girl's"
                        elif "KID" in prod_text:
                            excel_val = "Kid's"
                        else:
                            gender = detect_gender_from_sku(tag.get("SKU"))
                            if gender == "WOMENS":
                                excel_val = "Women's"
                            elif gender == "MENS":
                                excel_val = "Men's"
                            elif gender == "BOYS":
                                excel_val = "Boy's"
                            elif gender == "GIRLS":
                                excel_val = "Girl's"
                            elif gender == "KIDS":
                                excel_val = "Kid's"
                            else:
                                excel_val = None
            elif field_name == "Size":
                excel_sku = excel_row.get(sku_col) or pdf_sku_norm
                if not excel_sku or str(excel_sku).strip().upper() == "NAN" or str(excel_sku).strip() == "":
                    excel_sku = pdf_sku_norm
                _, _, extracted_size = extract_sku_details(excel_sku)
                excel_val = format_size_as_tag(extracted_size) if extracted_size else None
                if not pdf_val:
                    pdf_val = excel_val
            elif field_name == "Color":
                if tag_type == "B2B Bundle Sticker tag file":
                    pdf_val = tag.get("Description") or tag.get("Product")
                if not pdf_val or str(pdf_val).strip() == "" or str(pdf_val).strip().upper() == "NAN":
                    sku_to_use = tag.get("SKU") or excel_row.get(sku_col) or pdf_sku_norm
                    if sku_to_use:
                        _, extracted_color, _ = extract_sku_details(sku_to_use)
                        if extracted_color:
                            pdf_val = color_map.get(extracted_color, extracted_color)
                
                db_color_val = excel_row.get(excel_col) if excel_col else None
                if db_color_val and pd.notna(db_color_val) and str(db_color_val).strip() != "" and str(db_color_val).strip().upper() != "NAN":
                    excel_val = str(db_color_val).strip()
                else:
                    # Fetch color code from provided PDF SKU and look up in EAN rules Google Sheet
                    sku_for_color = str(tag.get("SKU") or pdf_sku_norm or (excel_row.get(sku_col) if excel_row is not None and sku_col else "") or "").strip()
                    extracted_color = None
                    if sku_for_color:
                        res = extract_sku_details_with_batch(sku_for_color)
                        if res and res[1]:
                            extracted_color = res[1]
                    if not extracted_color and pdf_sku_norm:
                        res = extract_sku_details_with_batch(pdf_sku_norm)
                        if res and res[1]:
                            extracted_color = res[1]
                    if extracted_color:
                        excel_val = gsheet_color_map.get(extracted_color) or color_map.get(extracted_color, extracted_color)
                    else:
                        excel_val = None
            elif field_name == "SKU":
                pdf_val = tag.get("SKU")
                excel_val = excel_row.get(excel_col) if excel_col else None
                if not excel_val or str(excel_val).strip().upper() == "NAN" or str(excel_val).strip() == "":
                    excel_val = pdf_sku_norm
            elif field_name == "EAN":
                pdf_val = tag.get("EAN") or tag.get("Barcode")
                excel_val = excel_row.get(barcode_col) if barcode_col else None
            else:
                if excel_col is None:
                    continue
                excel_val = excel_row.get(excel_col)

            pdf_norm = norm_fn(pdf_val)
            excel_norm = norm_fn(excel_val)

            if field_name == "Description":
                p_words = set(re.findall(r"\w+", str(pdf_norm).upper()))
                e_words = set(re.findall(r"\w+", str(excel_norm).upper()))
                common_words = p_words.intersection(e_words)
                has_conflict = check_conflicting_product_type(p_words, e_words)
                
                p_nouns = p_words.intersection(CLEAN_PRODUCT_NOUNS)
                e_nouns = e_words.intersection(CLEAN_PRODUCT_NOUNS)
                common_nouns = p_nouns.intersection(e_nouns)
                
                is_match = (
                    not has_conflict and (
                        pdf_norm == excel_norm
                        or len(common_words) >= 2
                        or bool(common_nouns)
                        or (bool(pdf_norm) and bool(excel_norm) and (
                            pdf_norm in excel_norm
                            or excel_norm in pdf_norm
                        ))
                    )
                )
                if not is_match:
                    g_desc = get_updated_description(tag.get("Style") or base_style_info, tag.get("SKU"), gsheet_dfs, tag_type=tag_type)
                    if g_desc:
                        g_norm = norm_fn(g_desc)
                        g_words = set(re.findall(r"\w+", str(g_norm).upper()))
                        common_g_words = p_words.intersection(g_words)
                        has_g_conflict = check_conflicting_product_type(p_words, g_words)
                        
                        pg_nouns = p_words.intersection(CLEAN_PRODUCT_NOUNS)
                        g_nouns = g_words.intersection(CLEAN_PRODUCT_NOUNS)
                        common_g_nouns = pg_nouns.intersection(g_nouns)
                        
                        if (
                            not has_g_conflict and (
                                pdf_norm == g_norm
                                or len(common_g_words) >= 2
                                or bool(common_g_nouns)
                                or (bool(pdf_norm) and bool(g_norm) and (
                                    pdf_norm in g_norm
                                    or g_norm in pdf_norm
                                ))
                            )
                        ):
                            is_match = True
                            excel_val = g_desc
                            excel_norm = g_norm
                status = "✅ Match" if is_match else "❌ Mismatch"
            elif field_name.startswith("Lot No"):
                def get_normalized_lot_parts(lot_str):
                    parts = str(lot_str).strip().upper().split("/")
                    base = parts[0].strip()
                    batch = parts[1].strip() if len(parts) > 1 else ""
                    batch_digits = "".join([c for c in batch if c.isdigit()])
                    if batch_digits:
                        try:
                            import re
                            match = re.match(r"^([A-Z]+)(0*)([0-9]+)$", batch)
                            if match:
                                batch = match.group(1) + match.group(3)
                            else:
                                batch = str(int(batch_digits))
                        except Exception:
                            pass
                    return base, batch

                p_base, p_batch = get_normalized_lot_parts(pdf_norm)
                e_base, e_batch = get_normalized_lot_parts(excel_norm)
                
                def clean_lot_base(base):
                    b = str(base).strip().upper()
                    two_letter = ("MT", "WT", "MS", "WS", "MV", "WV", "MI", "WI", "MJ", "WJ", "WP", "MP", "WB", "MB", "BT", "GP", "KD")
                    for tl in two_letter:
                        if b.startswith(tl):
                            b = b[len(tl):]
                            break
                    category_letters = {"O", "S", "P", "T", "M", "W", "K", "B", "G", "I", "J", "V", "D"}
                    while len(b) > 0 and b[0] in category_letters:
                        b = b[1:]
                    return b

                p_base_clean = clean_lot_base(p_base)
                e_base_clean = clean_lot_base(e_base)
                
                base_match = (p_base_clean == e_base_clean)
                if tag_type == "D2C Dress tag file" and field_name == "Lot No (GS1 Master)":
                    batch_match = True
                else:
                    batch_match = match_batch_code(p_batch, e_batch)
                
                is_match = base_match and batch_match
                status = "✅ Match" if is_match else "❌ Mismatch"
            elif field_name == "Fit":
                is_match = (
                    pdf_norm == excel_norm
                    or (bool(pdf_norm) and bool(excel_norm) and (
                        pdf_norm in excel_norm
                        or excel_norm in pdf_norm
                    ))
                )
                status = "✅ Match" if is_match else "❌ Mismatch"
            elif field_name == "Color":
                is_match = (
                    pdf_norm == excel_norm
                    or (bool(pdf_norm) and bool(excel_norm) and (
                        pdf_norm.startswith(excel_norm)
                        or excel_norm.startswith(pdf_norm)
                        or pdf_norm in excel_norm
                        or excel_norm in pdf_norm
                    ))
                )
                status = "✅ Match" if is_match else "❌ Mismatch"
            else:
                if field_name == "EAN" and is_simulated:
                    status = "❌ Not found in Excel"
                else:
                    status = "✅ Match" if pdf_norm == excel_norm else "❌ Mismatch"
            report_rows.append({
                "SKU": tag["SKU"],
                "Field": field_name,
                "PDF Value": pdf_val,
                "Excel Value": excel_val,
                "Status": status,
            })

    if not report_rows:
        return pd.DataFrame(columns=["SKU", "Field", "PDF Value", "Excel Value", "Status"])
    return pd.DataFrame(report_rows)


def main():
    if hasattr(sys.stdout, "reconfigure"):
        try:
            sys.stdout.reconfigure(encoding="utf-8")
        except Exception:
            pass

    ap = argparse.ArgumentParser(description="Compare dress-tag PDF against master Excel sheet")
    ap.add_argument("pdf", nargs="?", default=None)
    ap.add_argument("xlsx", nargs="?", default=None)
    ap.add_argument("--sheet", default=None)
    ap.add_argument("--out", default="tag_comparison_report.xlsx")
    args = ap.parse_args()

    script_dir = os.path.dirname(os.path.abspath(__file__))

    pdf_path = args.pdf
    if not pdf_path:
        pdfs = [f for f in os.listdir(script_dir) if f.lower().endswith(".pdf") and not f.startswith("~$")]
        if not pdfs:
            print("Error: No PDF file found in the script directory.", file=sys.stderr)
            sys.exit(1)
        pdf_path = os.path.join(script_dir, pdfs[0])
    elif not os.path.isabs(pdf_path):
        pdf_path = os.path.join(script_dir, pdf_path)

    xlsx_path = args.xlsx
    if not xlsx_path:
        xlsxs = [f for f in os.listdir(script_dir) if f.lower().endswith(".xlsx") and not f.startswith("~$") and not any(x in f.lower() for x in ["report", "google", "explore", "comparison"])]
        if not xlsxs:
            print("Error: No Excel (.xlsx) file found in the script directory.", file=sys.stderr)
            sys.exit(1)
        xlsx_path = os.path.join(script_dir, xlsxs[0])
    elif not os.path.isabs(xlsx_path):
        xlsx_path = os.path.join(script_dir, xlsx_path)

    out_path = args.out
    if not os.path.isabs(out_path):
        out_path = os.path.join(script_dir, out_path)

    print(f"Using PDF: {pdf_path}")
    print(f"Using Excel: {xlsx_path}")

    # Download updated MRP Google Sheet
    gsheet_dfs = {}
    gsheet_url = "https://docs.google.com/spreadsheets/d/1Q7nboN_Rezl807J0naA0QczTyoAQ6WM-KNmp_F26n5M/export?format=xlsx"
    gsheet_path = os.path.join(script_dir, "google_sheet_mrp.xlsx")
    print("\nDownloading updated MRP Google Sheet...")
    try:
        req = urllib.request.Request(gsheet_url, headers={'User-Agent': 'Mozilla/5.0'})
        with urllib.request.urlopen(req, timeout=15) as response:
            with open(gsheet_path, "wb") as f:
                f.write(response.read())
        print("Download successful.")
        xls = pd.ExcelFile(gsheet_path)
        for sheet_name in xls.sheet_names:
            gsheet_dfs[sheet_name] = pd.read_excel(xls, sheet_name=sheet_name)
    except Exception as e:
        print(f"Warning: Could not download/read updated MRP Google Sheet ({e}). Using local Excel MRP values as fallback.", file=sys.stderr)

    pdf_df = extract_pdf_tags(pdf_path)
    excel_df = extract_excel_master(xlsx_path, args.sheet)

    print(f"Extracted {len(pdf_df)} tags from PDF.")
    print(f"Extracted {len(excel_df)} rows from Excel.")

    if any(x in os.path.basename(pdf_path).lower() for x in ["bundle", "blank"]):
        tag_type = "B2B Bundle Sticker tag file"
    elif any(x in os.path.basename(pdf_path).lower() for x in ["b2b", "box", "sticker"]):
        tag_type = "B2B Box Sticker tag file"
    else:
        tag_type = "D2C Dress tag file"
    report_df = compare(pdf_df, excel_df, gsheet_dfs, tag_type=tag_type)

    n_mismatch = (report_df["Status"] != "✅ Match").sum()
    n_total = len(report_df)
    print(f"\n{n_total - n_mismatch}/{n_total} field checks passed.")
    if n_mismatch:
        print(f"{n_mismatch} issues found (showing first 50, see full report in Excel):")
        print(report_df[report_df["Status"] != "✅ Match"].head(50).to_string(index=False))

    try:
        with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
            pdf_df.to_excel(writer, sheet_name="PDF_Extracted", index=False)
            excel_df.to_excel(writer, sheet_name="Excel_Master", index=False)
            report_df.to_excel(writer, sheet_name="Comparison_Report", index=False)

        # Color the report sheet
        wb = openpyxl.load_workbook(out_path)
        ws = wb["Comparison_Report"]
        from openpyxl.styles import PatternFill
        green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        status_col_idx = report_df.columns.get_loc("Status") + 1
        for row_idx in range(2, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=status_col_idx)
            fill = green if "Match" in str(cell.value) and "Mis" not in str(cell.value) and "Not found" not in str(cell.value) else red
            for col_idx in range(1, ws.max_column + 1):
                ws.cell(row=row_idx, column=col_idx).fill = fill
        wb.save(out_path)

        print(f"\nReport saved to {out_path}")
    except PermissionError:
        print(f"\nERROR: Permission denied when writing to '{out_path}'.\n"
              "Please make sure the file is closed in Microsoft Excel or other programs and run the script again.", file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()
